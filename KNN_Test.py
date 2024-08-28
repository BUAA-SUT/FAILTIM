from KNN import *
import math
import numpy as np
import copy
import random
import json
import csv
from PublicFun import *
from openpyxl import load_workbook

random.seed(1)


def getSource(dynamic):
    Follow = []
    source_case_set = []
    for k in range(200):
        trainingSet = random.sample(dataset, int(len(dataset) * 0.6))
        testSet = copy.deepcopy(dataset)
        for i in range(len(trainingSet)):
            testSet.remove(trainingSet[i])  # 被测样本
        testSet = random.sample(testSet, 1)
        source_case_set.append([trainingSet, testSet])

    for i in range(len(source_case_set)):
        MGS, follow_case = MTG(source_case_set[i], dynamic)
        Follow.append(follow_case)

    MG = []
    for i in range(len(source_case_set)):
        MGS, Result = riskIndex(source_case_set[i], dynamic)
        MG.append(MGS)
    source_case_set = modifycase_fairness(source_case_set, MG)

    if len(source_case_set) == 0:
        print('测试用例集不满足要求')

    return source_case_set


def riskIndex(argv, dynamic):
    MGS = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        knn = KNN()
        knn.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_a = knn.getPredications()
        dynamic.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_m = dynamic.getPredications()
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)

    return MGS, Result


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/result1.xlsx'
    wb = load_workbook(path)
    string = 'KNN'
    del wb[string]
    ws = wb.create_sheet(string)
    with open('/Applications/work/data/MT/FAILTIM/Result/KNN/iris.data', 'rt') as csvfile:
        lines = csv.reader(csvfile)
        dataset = list(lines)
    source = []
    for mu in range(1, 11):
        dynamic = KNNFactory("MU_" + str(mu) + "_KNN").getKNN()
        source_case_set = getSource(dynamic)
        MG_set = []
        pf_set = []
        for i in range(len(source_case_set)):
            MG, pf = riskIndex(source_case_set[i], dynamic)
            MG_set.append(MG)
            pf_set.append(pf)

        if len(MG_set) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics(row, ws, mu, MG_set, pf_set)

    wb.save(path)

