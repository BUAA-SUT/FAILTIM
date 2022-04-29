# RQ1

from MATH import *
from openpyxl import load_workbook
from publicFun import *
import json

def getTestcase(mr_list, test_case, num_of_samples):
    # for i in range(num_of_samples):
    #     test_case.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
    #     test_case.generateRandomTestcase()
    # for i in range(num_of_samples):
    #     for j in range(len(mr_list)):
    #         mr = mr_list[j]
    #         mr.setTestCase(test_case)
    #         mr.original_ts.setInputOutput("input{}.fa".format(i), "reference{}.fa".format(i), "e{}.fa".format(i), "output{}.txt".format(i))
    #         mr.getFollowInput(j)
    for i in range(num_of_samples):
        test_case.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
        test_case.generateInput()  # 生成原始测试用例
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("input{}.txt".format(i),"output{}.txt".format(i))
            mr.getFollowInput(j)
            for k in range(len(mr_list)):
                mr = mr_list[k]
                mr.setTestCase(test_case)
                mr.original_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}.txt".format(i, j))
                mr.getFollowInput(k)

def getMG(mu, ts, num_of_samples, mr_list):
    MGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        original_output = MR().getResults(ts)
        followup_ts = ts
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts)
            expected_output = mr.getExpectedOutput(original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}.txt".format(i, m), "output{}_{}_{}.txt".format(mu, i, m))
            # ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
            original_output = MR().getResults(ts)
            followup_ts = ts
            for n in range(len(mr_list)):
                # followup_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}".format(i, j), "outtree_{}_{}".format(i, j))
                followup_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n), "output{}_{}_{}_{}.txt".format(mu, i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts)
                expected_output = mr.getExpectedOutput(original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
            MGs.append(MG)
        MGS.append(MGs)
    return MGS


def getpf(mu, ts, num_of_samples, mr_list):
    pf = []
    for i in range(num_of_samples):
        result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("input{}.txt".format(i),"output{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result[0] = 1

        for j in range(len(mr_list)):
            ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(0, i, j))
            original_output = MR().getResults(ts)
            program_ts = ts
            program_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
            program_output = MR().getResults(program_ts)
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result[j + 1] = 1

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
                                           "output{}_{}_{}_{}.txt".format(0, i, m, n))
                original_output = MR().getResults(ts)
                program_ts = ts
                program_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n), "output{}_{}_{}_{}.txt".format(mu, i, m, n))
                program_output = MR().getResults(program_ts)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
        pf.append(result)
    return pf

def riskIndex(MGS, Result):
    index = []
    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    # 统计全部的v和s
    sum_s = 0
    sum_v = 0
    for i in range(len(MGS)):
        sum_s += MGS[i].count(0)
        sum_v += MGS[i].count(1)
    # if not sum_s + sum_v == (len(MGS) * len(MGS[0])):
    #     print('异常')

    for i in range((len(MGS[0]) * (len(MGS[0]) + 1) + 1)):
        if i == 0:
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
        elif i <= len(MGS[0]):
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
            if MGS[0][i - 1] == 0:
                es += 1
                ns -= 1
            else:
                ev += 1
                nv -= 1
        else:
            t = int((i - 1) / len(MGS[0]))
            if MGS[t][(i - 1) % len(MGS[0])] == 0:
                es = 1
                ns = sum_s - es
                ev = 0
                nv = sum_v
            else:
                ev = 1
                nv = sum_v - ev
                es = 0
                ns = sum_s
        index.append([ev, es, nv, ns])
    return index

def getRisk(Index):
    Risk = []
    for i in range(len(Index)):
        risk = []
        index = Index[i]
        for j in range(len(index)):
            try:
                if index[j][0] + index[j][1] == 0:
                    formula = [-1000 for i in range(30)]
                elif index[j][0] + index[j][2] == 0:  # v = 0
                    formula = [-1000 for i in range(30)]
                elif index[j][1] + index[j][3] == 0:  # s = 0
                    formula = [1000 for i in range(30)]
                else:
                    formula = riskformula(index[j])
                risk.append(formula)
            except:
                print(index[j])
        Risk.append(risk)
    return Risk

def getIndex(MG, pf):
    Index = []
    for i in range(len(MG)):
        index = riskIndex(MG[i], pf[i])
        Index.append(index)
    MG, pf, Index = modifycase_fairness3(MG, pf, Index)
    return MG, pf, Index

def modifycase(MG, pf):
    index = []
    for i in range(len(pf)):
        if pf[i].count(1) > len(pf[i]) * 0.4:
            index.append(i)
    MG = [MG[i] for i in range(len(MG)) if (i not in index)]
    pf = [pf[i] for i in range(len(pf)) if (i not in index)]
    return MG, pf


if __name__ == "__main__":
    myenv = MyEnv()
    myenv.CreateWorkingDirs()
    path = '/Applications/work/data/MT/FAILTIM/Result/RQ3_3.xlsx'
    wb = load_workbook(path)
    string = 'MATH'
    del wb[string]
    ws = wb.create_sheet(string)
    row = 1
    ts = TestCase()
    num_of_samples = 100  # 测试用例个数
    mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9()]
    # getTestcase(mr_list, ts, num_of_samples)
    # MG = getMG(1, ts, num_of_samples, mr_list)
    MGSet = []
    pfset = []
    for mu in range(1, 14):
        # MG = getMG(mu, ts, num_of_samples, mr_list)
        # pf = getpf(mu, ts, num_of_samples, mr_list)
        #
        # MG, pf, Index = getIndex(MG, pf)
        # Risk = getRisk(Index)
        # # 存数据
        # data = {
        #     'pf': pf, 'MG': MG, 'Risk': Risk, 'Index': Index
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('./result/mutant' + str(mu) + '_rq3.json', 'w') as f:
        #     json.dump(json_str, f)
        # 读数据
        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']
        MGSet.append(MG)
        pfset.append(pf)
        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue
        row = getMetrics_rq3_3(row, ws, mu, MG, pf)

    wb.save(path)
