'''
RQ3
Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
from KNN import *
import math
import numpy as np
import copy
import random
import json
import csv
from publicFun import *
from openpyxl import load_workbook

random.seed(1)


def riskIndex(argv, dynamic):
    MGS = []
    index = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        knn = KNN()
        knn.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_a = knn.getPredications()
        dynamic.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_m = dynamic.getPredications()
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    # p or f统计完
    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3



    # 统计全部的v和s
    sum_s = 0
    sum_v = 0
    for i in range(len(MGS)):
        sum_s += MGS[i].count(0)
        sum_v += MGS[i].count(1)
    # if not sum_s + sum_v == (len(MGS) * len(MGS[0])):
    #     print('异常')

    for i in range(len(testcase)):
        if i == 0:
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
        elif i <= len(MGS[0]):
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
            if MGS[0][i-1] == 0:
                es += 1
                ns -= 1
            elif MGS[0][i-1] == 1:
                ev += 1
                nv -= 1
        else:
            t = int((i-1) / len(MGS[0]))
            if MGS[t][(i-1) % len(MGS[0])] == 0:
                es = 1
                ns = sum_s - es
                ev = 0
                nv = sum_v
            elif MGS[t][(i-1) % len(MGS[0])] == 1:
                ev = 1
                nv = sum_v - ev
                es = 0
                ns = sum_s
            else:
                es = 0
                ns = sum_s
                ev = 0
                nv = sum_v
        index.append([ev, es, nv, ns])
    return MGS, index, Result


def getRisk(source_case_set):
    MG = []
    pf = []
    Index = []
    for i in range(len(source_case_set)):
        MGS, index, Result = riskIndex(source_case_set[i], dynamic)  # MG里面有3
        MG.append(MGS)
        pf.append(Result)
        Index.append(index)

    Risk = []
    for i in range(len(Index)):
        risk = []
        index = Index[i]
        for j in range(len(index)):
            try:
                if index[j][0] + index[j][1] == 0:
                    formula = [-1000 for i in range(30)]
                elif index[j][0] + index[j][2] == 0:  # v = 0
                    formula = [-1000 for i in range(30)]
                elif index[j][1] + index[j][3] == 0:  # s = 0
                    formula = [1000 for i in range(30)]
                else:
                    formula = riskformula(index[j])
                risk.append(formula)
            except:
                print(index[j])
        Risk.append(risk)
    return pf, MG, Index, Risk


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/RQ3_3.xlsx'
    wb = load_workbook(path)
    string = 'knn'
    del wb[string]
    ws = wb.create_sheet(string)
    MGSet = []
    pfset = []
    for mu in range(1, 5):  # 6-10 全为等价mutant
        # dynamic = KNNFactory("MU_" + str(mu) + "_KNN").getKNN()
        # # 读测试用例
        # with open('Result/KNN/mutant' + str(mu) + '_rq1.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # source_case_set = data['source_case_set']
        #
        # pf, MG, Index, Risk = getRisk(source_case_set)
        # data = {
        #         'pf': pf, 'MG': MG, 'Risk': Risk, 'Index': Index
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('Result/KNN/mutant' + str(mu) + '_rq3.json', 'w') as f:
        #     json.dump(json_str, f)
        # 读数据
        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json',
                  'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']
        pfset.append(pf)
        MGSet.append(MG)

        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics_rq3_3(row, ws, mu, MG, pf)
    wb.save(path)



