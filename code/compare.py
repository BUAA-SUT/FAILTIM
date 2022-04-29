from openpyxl import load_workbook
import numpy as np
# wb1 = load_workbook('/Applications/work/data/MT/FAILTIM/Result/RQ1_old3.xlsx')
# wb2 = load_workbook('/Applications/work/data/MT/FAILTIM/Result/RQ1_new3.xlsx')
# path = '/Applications/work/data/MT/FAILTIM/Result/每个公式3.xlsx'
# wb3 = load_workbook(path)

Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice','Goodman','Tarantula',
           'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto','Hamming etc.', 'Euclid', 'Scott',
           'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

wb_r = load_workbook('/Applications/work/data/MT/FAILTIM/Result/RQ1_3.xlsx')
wb_w = load_workbook('/Applications/work/data/MT/FAILTIM/Result/RQ1_3_ave.xlsx')
program = ['trisquare', 'determinant', 'sparsemat', 'tcas', 'knn', 'printtokens','Num', 'MATH', 'DNA']
Ave = []
Z = []
for k in program:
    row = 1
    del wb_w[k]
    ws_w = wb_w.create_sheet(k)
    ws_r = wb_r[k]

    max_row = ws_r.max_row
    datadist = {}
    V_MGlist = []
    FFSlist = []
    p1list = []
    p2_1list = []
    p2_2list = []
    p3_1list = []
    p3_2list = []
    Failed = []
    for i in range(len(Formula)):
        V_MGlist.append([])
        FFSlist.append([])
        p1list.append([])
        p2_1list.append([])
        p2_2list.append([])
        p3_1list.append([])
        p3_2list.append([])
        Failed.append([])
    t = 0
    n = 2
    while True:
        p1list[t].append(ws_r.cell(n, 2).value)
        p2_1list[t].append(ws_r.cell(n, 3).value)
        p2_2list[t].append(ws_r.cell(n, 4).value)
        p3_1list[t].append(ws_r.cell(n, 5).value)
        p3_2list[t].append(ws_r.cell(n, 6).value)
        V_MGlist[t].append(ws_r.cell(n, 7).value)
        Failed[t].append(ws_r.cell(n, 8).value)
        FFSlist[t].append(ws_r.cell(n, 9).value)
        t = t + 1
        n = n + 1
        if t == 30:
            t = 0
            n = n + 3
        if n > max_row:
            break
    tablelist = {"Formulas": ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for i in range(len(Formula)):
        value = [round(np.mean(p1list[i]), 2), round(np.mean(p2_1list[i]), 2), round(np.mean(p2_2list[i]), 2),
                 round(np.mean(p3_1list[i]), 2), round(np.mean(p3_2list[i]), 2),
                 round(np.mean(V_MGlist[i]), 2),
                 round(np.mean(Failed[i]), 2), round(np.mean(FFSlist[i]), 2)]
        data = {
            Formula[i]: value
        }
        datadist.update(data)

    # 写入
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws_w.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws_w.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    wb_w.save('/Applications/work/data/MT/FAILTIM/Result/RQ1_3_ave.xlsx')

# program = ['MATH']
# 读
# 求平均
# 写
# Ave = []
# Z = []
# for k in program:
#     row = 1
#     # del wb3[k]
#     ws1 = wb3[k]
#     ws = wb1[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value + ws.cell(n, 5).value
#         # value = ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     tablelist = {"Formulas": 'old'}
#     datadist.update(tablelist)
#     for i in range(len(Formula)):
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     # 写入
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws1.cell(row, 1).value = i  # 添加第 1 列的数据
#         ws1.cell(row, 2).value = j
#         row += 1  # 行数
#     wb3.save(path)
#     # z = zip(datadist.values(), datadist.keys())
#     # 写入
#     # z = sorted(z, reverse=True)
#     # values = []
#     # for i in range(len(z)):
#     #     values.append(z[i][0])
#     # ave = round(np.mean(values), 2)
#     # Z.append(z)
#     # Ave.append(ave)
#
# for k in program:
#     row = 1
#     # del wb3[k]
#     ws1 = wb3[k]
#     ws = wb2[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value + ws.cell(n, 5).value
#         # value = ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     tablelist = {"Formulas": 'new'}
#     datadist.update(tablelist)
#     for i in range(len(Formula)):
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     # 写入
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws1.cell(row, 1).value = i  # 添加第 1 列的数据
#         ws1.cell(row, 3).value = j
#         row += 1  # 行数
#     wb3.save(path)

# wb1 = load_workbook('/Applications/work/data/MT/FAILTIM/Result/compare.xlsx')
# path = '/Applications/work/data/MT/FAILTIM/Result/compare_ave.xlsx'
# wb3 = load_workbook(path)
# for k in program:
#     row = 1
#     # del wb3[k]
#     ws1 = wb3[k]
#     ws = wb1[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value
#         # value = ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     tablelist = {"Formulas": 'equal'}
#     datadist.update(tablelist)
#     for i in range(len(Formula)):
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     # 写入
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws1.cell(row, 1).value = i  # 添加第 1 列的数据
#         ws1.cell(row, 2).value = j
#         row += 1  # 行数
#     wb3.save(path)
#
# for k in program:
#     row = 1
#     # del wb3[k]
#     ws1 = wb3[k]
#     ws = wb1[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 3).value
#         # value = ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     tablelist = {"Formulas": 'new'}
#     datadist.update(tablelist)
#     for i in range(len(Formula)):
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     # 写入
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws1.cell(row, 1).value = i  # 添加第 1 列的数据
#         ws1.cell(row, 3).value = j
#         row += 1  # 行数
#     wb3.save(path)
#
# for k in program:
#     row = 1
#     # del wb3[k]
#     ws1 = wb3[k]
#     ws = wb1[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 4).value
#         # value = ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     tablelist = {"Formulas": 'old'}
#     datadist.update(tablelist)
#     for i in range(len(Formula)):
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     # 写入
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws1.cell(row, 1).value = i  # 添加第 1 列的数据
#         ws1.cell(row, 4).value = j
#         row += 1  # 行数
#     wb3.save(path)