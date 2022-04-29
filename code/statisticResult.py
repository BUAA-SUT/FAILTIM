from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import ticker
import scipy.stats as stats
import json

Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
           'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
           'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
           'Arithmetic Mean', 'Cohen', 'Fleiss']

path = '/Applications/work/data/MT/FAILTIM/Result/correlation_with2.xlsx'
wb = load_workbook(path)
path2 = '/Applications/work/data/MT/FAILTIM/Result/correlation_without.xlsx'
wb2 = load_workbook(path2)
program = ['trisquare', 'determinant', 'sparsemat', 'tcas', 'knn', 'MATH', 'printtokens',  'Num', 'DNA']  #
# program = ['Num']
# for k in program:
#     del wb[k]
#     ws = wb.create_sheet(k)
#     row = 1
#     for mu in range(10, 14):
#         with open('/Applications/work/data/MT/FAILTIM/Result/' + k + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
#             data = json.load(load_f)
#         data = json.loads(data)
#         pf = data['pf']
#         MG = data['MG']
#
#         if len(MG) == 0:
#             print("Mutant" + str(mu) + '不符合要求')
#             continue
#
#         # 统计指标
#         datadist = {}
#         tablelist = {"Mutant" + str(mu): ['equal', 'new', 'old']}
#         datadist.update(tablelist)
#
#         with open('/Applications/work/data/MT/FAILTIM/Result/' + k + '/mutant' + str(mu) + '_rq1_newcases.json', 'r') as load_f:
#             data = json.load(load_f)
#         data = json.loads(data)
#         newcases = data['cases']
#
#         with open('/Applications/work/data/MT/FAILTIM/Result/' + k + '/mutant' + str(mu) + '_rq1_oldcases.json', 'r') as load_f:
#             data = json.load(load_f)
#         data = json.loads(data)
#         oldcases = data['cases']
#
#         for i in range(len(Formula)):
#             Equal = []
#             New = []
#             Old = []
#             for m in range(len(newcases[i])):
#                 equal = 0
#                 new = 0
#                 old = 0
#                 for n in range(len(newcases[i][m])):
#                     if newcases[i][m][n] == oldcases[i][m][n]:
#                         equal += 1
#                     elif newcases[i][m][n] == 1 and not oldcases[i][m][n] == 1:
#                         new += 1
#                     elif oldcases[i][m][n] == 1 and not newcases[i][m][n] == 1:
#                         old += 1
#                     else:
#                         equal += 1
#                 Equal.append(equal / len(newcases[i][m]))
#                 New.append(new / len(newcases[i][m]))
#                 Old.append(old / len(newcases[i][m]))
#
#             value = [round(np.mean(Equal) * 100, 2), round(np.mean(New) * 100, 2), round(np.mean(Old) * 100, 2)]
#             data = {
#                 Formula[i]: value
#             }
#             datadist.update(data)
#         for i, j in datadist.items():  # i--公式名称, j--指标值
#             ws.cell(row, 1).value = i  # 添加第 1 列的数据
#             for col in range(2, len(j) + 2):  # values列表中索引
#                 ws.cell(row, col).value = j[col - 2]
#             row += 1  # 行数
#         row += 2  # 行数
#
#     wb.save(path)


# program = ['trisquare']
# Data = {}
# Datalist = []
# for k in program:
#     ws = wb[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     for i in range(1, max_row + 1):
#         t = (i-1) % len(Formula)
#         value = ws.cell(i, 2).value + ws.cell(i, 3).value
#         Value[t].append(value)
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
#         #     continue
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     Datalist.append(datadist)
#
# sort = []
# for i in range(len(Datalist)):
#     z = zip(Datalist[i].values(), Datalist[i].keys())
#     z = sorted(z, reverse=True)
#     sort.append(z)
#
# for i in range(len(Formula)):
#     # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
#     #     continue
#     dd = []
#     for j in range(len(program)):
#         dd.append(Datalist[j][Formula[i]])
#     data = {
#         Formula[i]: round(np.mean(dd), 2)
#     }
#     Data.update(data)
# z = zip(Data.values(), Data.keys())
# z = sorted(z, reverse=True)
# values = []
# for i in range(len(z)):
#     values.append(z[i][0])
# ave = round(np.mean(values), 2)

# ws = wb['d_2']
# max_row = ws.max_row
# datadist = {}
# Value = []
# for i in range(len(Formula)):
#     Value.append([])
# for i in range(1, max_row + 1):
#     t = (i - 1) % len(Formula)
#     value = ws.cell(i, 2).value + ws.cell(i, 3).value
#     Value[t].append(value)
# for i in range(len(Formula)):
#     # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
#     #     continue
#     data = {
#         Formula[i]: round(np.mean(Value[i]), 2)
#     }
#     datadist.update(data)
# z = zip(datadist.values(), datadist.keys())
# z = sorted(z, reverse=True)
# values = []
# for i in range(len(z)):
#     values.append(z[i][0])
# ave = round(np.mean(values), 2)

# program = ['trisquare_y', 'determinant_y', 'sparsemat_y', 'tcas_y', 'knn_y', 'M_y', 'P_y',  'N_y', 'DNA_y',
#            'trisquare_n', 'determinant_n', 'sparsemat_n', 'tcas_n', 'knn_n', 'M_n', 'P_n',  'N_n', 'DNA_n']
# program = ['trisquare', 'determinant', 'sparsemat', 'tcas', 'knn', 'MATH', 'printtokens',  'Num', 'DNA']
# program = ['printtokens']
# d = 10
# # program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'k_' + str(d),
# #            'M_' + str(d), 'P_' + str(d), 'N_' + str(d), 'Dn_' + str(d)]
# program = [
#            'M_' + str(d), 'Dn_' + str(d)]
# # program = ['Dn_' + str(0), 'Dn_' + str(1), 'Dn_' + str(2), 'Dn_' + str(3), 'Dn_' + str(4), 'Dn_' + str(5),
# #            'Dn_' + str(6), 'Dn_' + str(7), 'Dn_' + str(8), 'Dn_' + str(9), 'Dn_' + str(10)]
# strr = 'tr_'
# program = [strr + str(0), strr + str(1), strr + str(2), strr + str(3), strr + str(4), strr + str(5), strr + str(6), strr + str(7), strr + str(8)
#            ]
# # # program = ['Dn_' + str(9), 'Dn_' + str(10)], strr + str(6), strr + str(7), strr + str(8), strr + str(9), strr + str(10)
# Ave = []
# Z = []
# for k in program:
#     ws = wb[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#         #     continue
#         # if not Formula[i] == 'Arithmetic Mean':
#         #     continue
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     z = zip(datadist.values(), datadist.keys())
#     z = sorted(z, reverse=True)
#     values = []
#     for i in range(len(z)):
#         values.append(z[i][0])
#     ave = round(np.mean(values), 2)
#     Z.append(z)
#     Ave.append(ave)

# # Ave2 = []
# # for k in program:
# #     ws = wb[k]
# #     max_row = ws.max_row
# #     datadist = {}
# #     Value = []
# #     for i in range(len(Formula)):
# #         Value.append([])
# #     t = 0
# #     n = 2
# #     while True:
# #         value = ws.cell(n, 8).value
# #         Value[t].append(value)
# #         t = t + 1
# #         n = n + 1
# #         if t == 30:
# #             t = 0
# #             n = n + 3
# #         if n > max_row:
# #             break
# #     for i in range(len(Formula)):
# #         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #         #     continue
# #         # if not Formula[i] == 'Russel&Rao':
# #         #     continue
# #         data = {
# #             Formula[i]: round(np.mean(Value[i]), 2)
# #         }
# #         datadist.update(data)
# #     z = zip(datadist.values(), datadist.keys())
# #     z = sorted(z, reverse=True)
# #     values = []
# #     for i in range(len(z)):
# #         values.append(z[i][0])
# #     ave = round(np.mean(values), 2)
# #     Ave2.append(ave)
#
# # #
# # program = ['trisquare', 'determinant', 'sparsemat', 'tcas', 'knn', 'MATH', 'printtokens',  'Num', 'DNA']
# # program = ['trisquare_n', 'determinant_n', 'sparsemat_n', 'tcas_n', 'knn_n',  'DNA_n', 'M_n', 'P_n',  'N_n']
# # program = ['trisquare']
# # program = ['trisquare_y', 'determinant_y', 'sparsemat_y', 'tcas_y', 'knn_y',  'DNA_y', 'M_y', 'P_y',  'N_y']
# # d = 10
# # program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'Dn_' + str(d),
# #            'k_' + str(d), 'M_' + str(d), 'P_' + str(d), 'N_' + str(d)]
# # program = ['Dn_' + str(d)]
# # program = ['Dn_' + str(0), 'Dn_' + str(1), 'Dn_' + str(2), 'Dn_' + str(3), 'Dn_' + str(4), 'Dn_' + str(5),
# #            'Dn_' + str(6), 'Dn_' + str(7), 'Dn_' + str(8), 'Dn_' + str(9), 'Dn_' + str(10)]
# # strr = 'tr_'
# # # program = [strr + str(0), strr + str(1), strr + str(2), strr + str(3), strr + str(4), strr + str(5),
# # #            strr + str(6), strr + str(7), strr + str(8)]
# # program = [strr + str(7)]
# # # 方法一，分两步平均
# # Data = {}
# # Datalist = []
# # for k in program:
# #     ws = wb[k]
# #     max_row = ws.max_row
# #     datadist = {}
# #     Value = []
# #     for i in range(len(Formula)):
# #         Value.append([])
# #     t = 0
# #     n = 2
# #     while True:
# #         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
# #         Value[t].append(value)
# #         t = t + 1
# #         n = n + 1
# #         if t == 30:
# #             t = 0
# #             n = n + 3
# #         if n > max_row:
# #             break
# #     for i in range(len(Formula)):
# #         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #         #     continue
# #         # if not Formula[i] == 'Russel&Rao':
# #         #     continue
# #         data = {
# #             Formula[i]: round(np.mean(Value[i]), 2)
# #         }
# #         datadist.update(data)
# #     Datalist.append(datadist)
# #
# # # sort = []
# # # for i in range(len(Datalist)):
# # #     z = zip(Datalist[i].values(), Datalist[i].keys())
# # #     z = sorted(z, reverse=True)
# # #     sort.append(z)
# #
# # for i in range(len(Formula)):
# #     # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #     #     continue
# #     # if not Formula[i] == 'Russel&Rao':
# #     #     continue
# #     dd = []
# #     for j in range(len(program)):
# #         dd.append(Datalist[j][Formula[i]])
# #     data = {
# #         Formula[i]: round(np.mean(dd), 2)
# #     }
# #     Data.update(data)
# # z = zip(Data.values(), Data.keys())
# # z = sorted(z, reverse=True)
# # values = []
# # for i in range(len(z)):
# #     values.append(z[i][0])
# # ave = round(np.mean(values), 2)
#
#
# # program = ['trisquare']
# # program = ['trisquare_n', 'determinant_n', 'sparsemat_n', 'tcas_n', 'knn_n',  'DNA_n', 'M_n', 'P_n',  'N_n']
# # program = ['trisquare_n']
# # program = ['trisquare_y', 'determinant_y', 'sparsemat_y', 'tcas_y', 'knn_y',  'DNA_y', 'M_y', 'P_y',  'N_y']
# # d = 10
# # # program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'k_' + str(d),
# # #            'M_' + str(d), 'P_' + str(d), 'N_' + str(d), 'Dn_' + str(d)]
# # program = [
# #            'M_' + str(d), 'Dn_' + str(d)]
# # # # program = ['Dn_' + str(0), 'Dn_' + str(1), 'Dn_' + str(2), 'Dn_' + str(3), 'Dn_' + str(4), 'Dn_' + str(5),
# # # #            'Dn_' + str(6), 'Dn_' + str(7), 'Dn_' + str(8), 'Dn_' + str(9), 'Dn_' + str(10)]
# # # strr = 'tr_'
# # # # program = [strr + str(0), strr + str(1), strr + str(2), strr + str(3), strr + str(4), strr + str(5),
# # # #            strr + str(6), strr + str(7), strr + str(8)]
# # # program = [strr + str(0)]
# # # Data = {}
# # # Datalist = []
# # # for k in program:
# # #     ws = wb[k]
# # #     max_row = ws.max_row
# # #     datadist = {}
# # #     Value = []
# # #     for i in range(len(Formula)):
# # #         Value.append([])
# # #     t = 0
# # #     n = 2
# # #     while True:
# # #         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
# # #         Value[t].append(value)
# # #         t = t + 1
# # #         n = n + 1
# # #         if t == 30:
# # #             t = 0
# # #             n = n + 3
# # #         if n > max_row:
# # #             break
# # #     for i in range(len(Formula)):
# # #         # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
# # #         #     continue
# # #         data = {
# # #             Formula[i]: round(np.mean(Value[i]), 2)
# # #         }
# # #         datadist.update(data)
# # #     Datalist.append(datadist)
# # #
# # # sort = []
# # # for i in range(len(Datalist)):
# # #     z = zip(Datalist[i].values(), Datalist[i].keys())
# # #     z = sorted(z, reverse=True)
# # #     sort.append(z)
# # #
# # # for i in range(len(Formula)):
# # #     # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
# # #     #     continue
# # #     dd = []
# # #     for j in range(len(program)):
# # #         dd.append(Datalist[j][Formula[i]])
# # #     data = {
# # #         Formula[i]: round(np.mean(dd), 2)
# # #     }
# # #     Data.update(data)
# # # z = zip(Data.values(), Data.keys())
# # # z = sorted(z, reverse=True)
# # # values = []
# # # for i in range(len(z)):
# # #     values.append(z[i][0])
# # # ave = round(np.mean(values), 2)
# #
# #
# # # 方法二，先全部求和再平均
# # Data = {}
# # Datalist = []
# # for k in program:
# #     ws = wb[k]
# #     max_row = ws.max_row
# #     datadist = {}
# #     Value = []
# #     for i in range(len(Formula)):
# #         Value.append([])
# #     t = 0
# #     n = 2
# #     while True:
# #         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
# #         Value[t].append(value)
# #         t = t + 1
# #         n = n + 1
# #         if t == 30:
# #             t = 0
# #             n = n + 3
# #         if n > max_row:
# #             break
# #     for i in range(len(Formula)):
# #         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #         #     continue
# #         data = {
# #             Formula[i]: Value[i]
# #         }
# #         datadist.update(data)
# #     Datalist.append(datadist)
# #
# # for i in range(len(Formula)):
# #     # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #     #     continue
# #     dd = []
# #     value = []
# #     for j in range(len(program)):
# #         dd += Datalist[j][Formula[i]]
# #     data = {
# #         Formula[i]: dd
# #     }
# #     Data.update(data)
# #
# # withfs = zip(Data.values(), Data.keys())
# # withfs = list(withfs)
# #
# # value = []
# # for i in range(len(withfs)):
# #     value.append(np.mean(withfs[i][0]))
# # value_sort = sorted(value, reverse=True)
# # value_ave = round(np.mean(value), 2)
# # value_std = round(np.std(value), 2)
#
# Pairs = ["d≥1", "d≥2", "d≥3", "d≥4", "d≥5", "d≥6", "d≥7", "d≥8", "d≥9", "d≥10", "d≥11"]
#
# Data2 = {}
# for d in range(11):
#     program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'k_' + str(d),
#                'M_' + str(d), 'P_' + str(d), 'N_' + str(d), 'Dn_' + str(d)]
#     if 6 <= d <= 8:
#         program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d),
#                    'M_' + str(d), 'P_' + str(d), 'Dn_' + str(d)]
#     if d == 9:
#         program = ['d_' + str(d),
#                    'M_' + str(d), 'Dn_' + str(d)]
#     if d == 10:
#         program = [
#                    'M_' + str(d), 'Dn_' + str(d)]
# # # program = ['Dn_' + str(0), 'Dn_' + str(1), 'Dn_' + str(2), 'Dn_' + str(3), 'Dn_' + str(4), 'Dn_' + str(5),
# # #            'Dn_' + str(6), 'Dn_' + str(7), 'Dn_' + str(8), 'Dn_' + str(9), 'Dn_' + str(10)]
# # strr = 'tr_'
# # # program = [strr + str(0), strr + str(1), strr + str(2), strr + str(3), strr + str(4), strr + str(5),
# # #            strr + str(6), strr + str(7), strr + str(8)]
# # program = [strr + str(0)]
# # Data = {}
# # Datalist = []
# # for k in program:
# #     ws = wb[k]
# #     max_row = ws.max_row
# #     datadist = {}
# #     Value = []
# #     for i in range(len(Formula)):
# #         Value.append([])
# #     t = 0
# #     n = 2
# #     while True:
# #         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
# #         Value[t].append(value)
# #         t = t + 1
# #         n = n + 1
# #         if t == 30:
# #             t = 0
# #             n = n + 3
# #         if n > max_row:
# #             break
# #     for i in range(len(Formula)):
# #         # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
# #         #     continue
# #         data = {
# #             Formula[i]: round(np.mean(Value[i]), 2)
# #         }
# #         datadist.update(data)
# #     Datalist.append(datadist)
# #
# # sort = []
# # for i in range(len(Datalist)):
# #     z = zip(Datalist[i].values(), Datalist[i].keys())
# #     z = sorted(z, reverse=True)
# #     sort.append(z)
# #
# # for i in range(len(Formula)):
# #     # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
# #     #     continue
# #     dd = []
# #     for j in range(len(program)):
# #         dd.append(Datalist[j][Formula[i]])
# #     data = {
# #         Formula[i]: round(np.mean(dd), 2)
# #     }
# #     Data.update(data)
# # z = zip(Data.values(), Data.keys())
# # z = sorted(z, reverse=True)
# # values = []
# # for i in range(len(z)):
# #     values.append(z[i][0])
# # ave = round(np.mean(values), 2)
#
#
# # 方法二，先全部求和再平均
#     Data = {}
#     Datalist = []
#     for k in program:
#         ws = wb[k]
#         max_row = ws.max_row
#         datadist = {}
#         Value = []
#         for i in range(len(Formula)):
#             Value.append([])
#         t = 0
#         n = 2
#         while True:
#             value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
#             Value[t].append(value)
#             t = t + 1
#             n = n + 1
#             if t == 30:
#                 t = 0
#                 n = n + 3
#             if n > max_row:
#                 break
#         for i in range(len(Formula)):
#             # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#             #     continue
#             data = {
#                 Formula[i]: Value[i]
#             }
#             datadist.update(data)
#         Datalist.append(datadist)
#
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#         #     continue
#         dd = []
#         value = []
#         for j in range(len(program)):
#             dd += Datalist[j][Formula[i]]
#         data = {
#             Formula[i]: dd
#         }
#         Data.update(data)
#
#     withfs = zip(Data.values(), Data.keys())
#     withfs = list(withfs)
#
#     value = []
#     for i in range(len(withfs)):
#         value.append(np.mean(withfs[i][0]))
#
#     data2 = {
#         Pairs[d]: value
#     }
#     Data2.update(data2)
#
# # withfs = zip(Data.values(), Data.keys())
# # withfs = list(withfs)
# #
# # value = []
# # for i in range(len(withfs)):
# #     value.append(np.mean(withfs[i][0]))
# # value_sort = sorted(value, reverse=True)
# # value_ave = round(np.mean(value), 2)
# # value_std = round(np.std(value), 2)
#
#
# path2 = '/Applications/work/data/MT/FAILTIM/Result/t-test-rq4.xlsx'
# wb = load_workbook(path2)
# string = 'Sheet2'
# del wb[string]
# ws = wb.create_sheet(string)
# row = 1
# datadist = {}
# tablelist = {'pairs': ['t-score', 'degree of freedom']}
# datadist.update(tablelist)
# t_score = []
# fp = []
# mm = 55
# for m in Pairs:
#     for n in Pairs:
#         s = 0
#         if m == n:
#             continue
#         if not fp:
#             pass
#         else:
#             for k in fp:
#                 if m in k and n in k:
#                     s = 1
#                     break
#         if s == 1:
#             continue
#         string = m + '-' + n
#         fp.append(string)
#         sp2 = 0
#         x1 = Data2[m]
#         x2 = Data2[n]
#         for i in range(len(x1)):
#             sp2 += ((x1[i] - np.mean(x1)) ** 2 + (x2[i] - np.mean(x2)) ** 2)
#         sp2 /= (2 * len(x1) - 2)
#         t = abs(np.mean(x1) - np.mean(x2)) / np.sqrt(2 * sp2 / len(x1))
#         t_score.append(t)
#
#         value = [round(t, 4), 29]
#         mm -= 1
#         data = {
#             string: value
#         }
#         datadist.update(data)
#
# for i, j in datadist.items():  # i--公式名称, j--指标值
#     ws.cell(row, 1).value = i  # 添加第 1 列的数据
#     for col in range(2, len(j) + 2):  # values列表中索引
#         ws.cell(row, col).value = j[col - 2]
#     row += 1  # 行数
# wb.save(path2)

#####
#  每个样本表示mutant对所有公式求平均值，Mann-Whitney-U 检验
#####

# Ave2 = []
# for k in program:
#     ws = wb[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 8).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#         #     continue
#         # if not Formula[i] == 'Russel&Rao':
#         #     continue
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     z = zip(datadist.values(), datadist.keys())
#     z = sorted(z, reverse=True)
#     values = []
#     for i in range(len(z)):
#         values.append(z[i][0])
#     ave = round(np.mean(values), 2)
#     Ave2.append(ave)

# #
# program = ['trisquare', 'determinant', 'sparsemat', 'tcas', 'knn', 'MATH', 'printtokens',  'Num', 'DNA']
# program = ['trisquare_n', 'determinant_n', 'sparsemat_n', 'tcas_n', 'knn_n',  'DNA_n', 'M_n', 'P_n',  'N_n']
# program = ['trisquare']
# program = ['trisquare_y', 'determinant_y', 'sparsemat_y', 'tcas_y', 'knn_y',  'DNA_y', 'M_y', 'P_y',  'N_y']
# d = 10
# program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'Dn_' + str(d),
#            'k_' + str(d), 'M_' + str(d), 'P_' + str(d), 'N_' + str(d)]
# program = ['Dn_' + str(d)]
# program = ['Dn_' + str(0), 'Dn_' + str(1), 'Dn_' + str(2), 'Dn_' + str(3), 'Dn_' + str(4), 'Dn_' + str(5),
#            'Dn_' + str(6), 'Dn_' + str(7), 'Dn_' + str(8), 'Dn_' + str(9), 'Dn_' + str(10)]
# strr = 'tr_'
# # program = [strr + str(0), strr + str(1), strr + str(2), strr + str(3), strr + str(4), strr + str(5),
# #            strr + str(6), strr + str(7), strr + str(8)]
# program = [strr + str(7)]
# # 方法一，分两步平均
# Data = {}
# Datalist = []
# for k in program:
#     ws = wb[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#         #     continue
#         # if not Formula[i] == 'Russel&Rao':
#         #     continue
#         data = {
#             Formula[i]: round(np.mean(Value[i]), 2)
#         }
#         datadist.update(data)
#     Datalist.append(datadist)
#
# # sort = []
# # for i in range(len(Datalist)):
# #     z = zip(Datalist[i].values(), Datalist[i].keys())
# #     z = sorted(z, reverse=True)
# #     sort.append(z)
#
# for i in range(len(Formula)):
#     # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#     #     continue
#     # if not Formula[i] == 'Russel&Rao':
#     #     continue
#     dd = []
#     for j in range(len(program)):
#         dd.append(Datalist[j][Formula[i]])
#     data = {
#         Formula[i]: round(np.mean(dd), 2)
#     }
#     Data.update(data)
# z = zip(Data.values(), Data.keys())
# z = sorted(z, reverse=True)
# values = []
# for i in range(len(z)):
#     values.append(z[i][0])
# ave = round(np.mean(values), 2)
#
# Pairs = ["d≥1", "d≥2", "d≥3", "d≥4", "d≥5", "d≥6", "d≥7", "d≥8", "d≥9", "d≥10", "d≥11"]
# #
# # Data2 = {}
# # for d in range(11):
# #     program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'k_' + str(d),
# #                'M_' + str(d), 'P_' + str(d), 'N_' + str(d), 'Dn_' + str(d)]
# #     if 6 <= d <= 8:
# #         program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d),
# #                    'M_' + str(d), 'P_' + str(d), 'Dn_' + str(d)]
# #     if d == 9:
# #         program = ['d_' + str(d),
# #                    'M_' + str(d), 'Dn_' + str(d)]
# #     if d == 10:
# #         program = [
# #                    'M_' + str(d), 'Dn_' + str(d)]
# # # # program = ['Dn_' + str(0), 'Dn_' + str(1), 'Dn_' + str(2), 'Dn_' + str(3), 'Dn_' + str(4), 'Dn_' + str(5),
# # # #            'Dn_' + str(6), 'Dn_' + str(7), 'Dn_' + str(8), 'Dn_' + str(9), 'Dn_' + str(10)]
# # # strr = 'tr_'
# # # # program = [strr + str(0), strr + str(1), strr + str(2), strr + str(3), strr + str(4), strr + str(5),
# # # #            strr + str(6), strr + str(7), strr + str(8)]
# # # program = [strr + str(0)]
# # # Data = {}
# # # Datalist = []
# # # for k in program:
# # #     ws = wb[k]
# # #     max_row = ws.max_row
# # #     datadist = {}
# # #     Value = []
# # #     for i in range(len(Formula)):
# # #         Value.append([])
# # #     t = 0
# # #     n = 2
# # #     while True:
# # #         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
# # #         Value[t].append(value)
# # #         t = t + 1
# # #         n = n + 1
# # #         if t == 30:
# # #             t = 0
# # #             n = n + 3
# # #         if n > max_row:
# # #             break
# # #     for i in range(len(Formula)):
# # #         # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
# # #         #     continue
# # #         data = {
# # #             Formula[i]: round(np.mean(Value[i]), 2)
# # #         }
# # #         datadist.update(data)
# # #     Datalist.append(datadist)
# # #
# # # sort = []
# # # for i in range(len(Datalist)):
# # #     z = zip(Datalist[i].values(), Datalist[i].keys())
# # #     z = sorted(z, reverse=True)
# # #     sort.append(z)
# # #
# # # for i in range(len(Formula)):
# # #     # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
# # #     #     continue
# # #     dd = []
# # #     for j in range(len(program)):
# # #         dd.append(Datalist[j][Formula[i]])
# # #     data = {
# # #         Formula[i]: round(np.mean(dd), 2)
# # #     }
# # #     Data.update(data)
# # # z = zip(Data.values(), Data.keys())
# # # z = sorted(z, reverse=True)
# # # values = []
# # # for i in range(len(z)):
# # #     values.append(z[i][0])
# # # ave = round(np.mean(values), 2)
# #
# #
# # # 方法二，先全部求和再平均
# #     Data = {}
# #     Datalist = []
# #     for k in program:
# #         ws = wb[k]
# #         max_row = ws.max_row
# #         datadist = {}
# #         Value = []
# #         for i in range(len(Formula)):
# #             Value.append([])
# #         t = 0
# #         n = 2
# #         while True:
# #             value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
# #             Value[t].append(value)
# #             t = t + 1
# #             n = n + 1
# #             if t == 30:
# #                 t = 0
# #                 n = n + 3
# #             if n > max_row:
# #                 break
# #         for i in range(len(Formula)):
# #             # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #             #     continue
# #             data = {
# #                 Formula[i]: Value[i]
# #             }
# #             datadist.update(data)
# #         Datalist.append(datadist)
# #
# #     for i in range(len(Formula)):
# #         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
# #         #     continue
# #         dd = []
# #         value = []
# #         for j in range(len(program)):
# #             dd += Datalist[j][Formula[i]]
# #         data = {
# #             Formula[i]: dd
# #         }
# #         Data.update(data)
# #
# #     withfs = zip(Data.values(), Data.keys())
# #     withfs = list(withfs)
# #
# #     value = []
# #     for i in range(len(withfs)):
# #         value.append(np.mean(withfs[i][0]))
# #
# #     data2 = {
# #         Pairs[d]: value
# #     }
# #     Data2.update(data2)
#
# Data2 = {}
# Datalist = []
# for d in range(11):
#     program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'k_' + str(d),
#                'M_' + str(d), 'P_' + str(d), 'N_' + str(d), 'Dn_' + str(d)]
#     if 6 <= d <= 8:
#         program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d),
#                    'M_' + str(d), 'P_' + str(d), 'Dn_' + str(d)]
#     if d == 9:
#         program = ['d_' + str(d),
#                    'M_' + str(d), 'Dn_' + str(d)]
#     if d == 10:
#         program = [
#                    'M_' + str(d), 'Dn_' + str(d)]
#     Data = {}
#     for k in program:
#         ws = wb[k]
#         max_row = ws.max_row
#         datadist = {}
#         Value = []
#         Name = []
#         for i in range(len(Formula)):
#             Value.append([])
#         t = 0
#         n = 2
#         Name.append(ws.cell(n-1, 1).value)
#         while True:
#             value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
#             Value[t].append(value)
#             t = t + 1
#             n = n + 1
#             if t == 30:
#                 t = 0
#                 n = n + 3
#                 if ws.cell(n - 1, 1).value is None:
#                     pass
#                 else:
#                     Name.append(ws.cell(n - 1, 1).value)
#             if n > max_row:
#                 break
#
#         for i in range(len(Name)):
#             # if Formula[i] == 'Naish1' or Formula[i] == 'Binary':
#             #     continue
#             dd = []
#             for j in range(len(Formula)):
#                 dd.append(Value[j][i])
#             data = {
#                 Name[i]: round(np.mean(dd), 2)
#             }
#             datadist.update(data)
#         data = {
#             k: datadist
#         }
#         Data.update(data)
#     Datalist.append(Data)
#
# path2 = '/Applications/work/data/MT/FAILTIM/Result/Mann-Whitney-U.xlsx'
# wb2 = load_workbook(path2)
# string = 'Sheet1'
# del wb2[string]
# ws2 = wb2.create_sheet(string)
# row = 1
# datadist = {}
# tablelist = {'case pairs': ['statistic', 'pvalue']}
# datadist.update(tablelist)
# t_score = []
# fp = []
# mm = 55
# for m in Pairs:
#     for n in Pairs:
#         s = 0
#         if m == n:
#             continue
#         if not fp:
#             pass
#         else:
#             for k in fp:
#                 if m in k and n in k:
#                     s = 1
#                     break
#         if s == 1:
#             continue
#         string = m + '-' + n
#         fp.append(string)
#         # 求x和y
#         # m永远小于n
#         if m == 'd≥10' or m == 'd≥11':
#             md = int(m[-2:]) - 1
#         else:
#             md = int(m[-1]) - 1
#         if n == 'd≥10' or n == 'd≥11':
#             nd = int(n[-2:]) - 1
#         else:
#             nd = int(n[-1]) - 1
#         Datax = Datalist[md]
#         Datay = Datalist[nd]
#         # 以Datay为基础找
#         x = []
#         y = []
#         d = nd
#         if d < 6:
#             program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d), 'k_' + str(d),
#                        'M_' + str(d), 'P_' + str(d), 'N_' + str(d), 'Dn_' + str(d)]
#         if 6 <= d <= 8:
#             program = ['tr_' + str(d), 'd_' + str(d), 's_' + str(d), 'tc_' + str(d),
#                        'M_' + str(d), 'P_' + str(d), 'Dn_' + str(d)]
#         if d == 9:
#             program = ['d_' + str(d), 'M_' + str(d), 'Dn_' + str(d)]
#         if d == 10:
#             program = ['M_' + str(d), 'Dn_' + str(d)]
#
#         for i in program:
#             if d == 10:
#                 ii = i[:-2] + str(md)
#             else:
#                 ii = i[:-1] + str(md)
#             key = list(Datay[i].keys())
#             for j in key:
#                 y.append(Datay[i][j])
#                 x.append(Datax[ii][j])
#
#         statistic, pvalue = stats.mannwhitneyu(x, y, alternative='two-sided')
#
#         value = [round(statistic, 4), round(pvalue, 4)]
#         data = {
#             string: value
#         }
#         datadist.update(data)
#
# for i, j in datadist.items():  # i--公式名称, j--指标值
#     ws2.cell(row, 1).value = i  # 添加第 1 列的数据
#     for col in range(2, len(j) + 2):  # values列表中索引
#         ws2.cell(row, col).value = j[col - 2]
#     row += 1  # 行数
# wb2.save(path2)


Data = {}
Datalist = []
Datafs = {}
Datalistfs = []
Datano = {}
Datalistno = []
for k in program:
    ws = wb[k]
    ws2 = wb2[k]
    max_row = ws.max_row
    max_col = ws.max_column
    datadist = {}
    Value = []
    datadistfs = {}
    Valuefs = []
    datadistno = {}
    Valueno = []
    for i in range(len(Formula)):
        Value.append([])
        Valuefs.append([])
        Valueno.append([])
    t = 0
    n = 2
    while True:
        value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
        Value[t].append(value)
        valuefs = ws.cell(n, max_col).value
        Valuefs[t].append(valuefs)
        valueno= ws2.cell(n, 2).value + ws2.cell(n, 4).value + ws2.cell(n, 6).value + 0.5 * ws2.cell(n, 3).value
        Valueno[t].append(valueno)
        t = t + 1
        n = n + 1
        if t == 30:
            t = 0
            n = n + 3
        if n > max_row:
            break
    for i in range(len(Formula)):
        # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
        #     continue
        data = {
            Formula[i]: Value[i]
        }
        datadist.update(data)
        datafs = {
            Formula[i]: Valuefs[i]
        }
        datadistfs.update(datafs)
        datano = {
            Formula[i]: Valueno[i]
        }
        datadistno.update(datano)
    Datalist.append(datadist)
    Datalistfs.append(datadistfs)
    Datalistno.append(datadistno)

for i in range(len(Formula)):
    # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
    #     continue
    dd = []
    value = []
    ddfs = []
    valuefs = []
    ddno = []
    valueno = []
    for j in range(len(program)):
        dd += Datalist[j][Formula[i]]
        ddfs += Datalistfs[j][Formula[i]]
        ddno += Datalistno[j][Formula[i]]
    data = {
        Formula[i]: dd
    }
    datafs = {
        Formula[i]: ddfs
    }
    datano = {
        Formula[i]: ddno
    }
    Data.update(data)
    Datafs.update(datafs)
    Datano.update(datano)

withfs = zip(Data.values(), Data.keys())
withfs = list(withfs)
withoutfs = zip(Datano.values(), Datano.keys())
withoutfs = list(withoutfs)
fs = zip(Datafs.values(), Datafs.keys())
fs = list(fs)

acc_with = []
acc_without = []
value_fs = []
for i in range(len(withfs[0][0])):
    vv = []
    vs = []
    vw = []
    for j in range(len(withfs)):
        vv.append(withfs[j][0][i])
        vw.append(withoutfs[j][0][i])
        vs.append(fs[j][0][i])
    acc_with.append(np.mean(vv))
    acc_without.append(np.mean(vw))
    value_fs.append(np.mean(vs) / 100)
improve = [(acc_without[i] - acc_with[i]) / acc_with[i] for i in range(0, len(acc_with))]
# improve = [(acc_without[i] - acc_with[i]) for i in range(0, len(acc_with))]

# 拟合曲线

points_tulpe = zip(value_fs, improve)  # [21:32]
z_points_tulpe = list(zip(value_fs, improve))
z = sorted(points_tulpe)
z = list(z)
x = []
y = []
for i in range(len(z)):
    x.append(z[i][0])
    y.append(z[i][1])
# plt.figure()
parameter = np.polyfit(x, y, 3)
y2 = []
for i in x:
    y2.append(parameter[0] * i ** 3 + parameter[1] * i ** 2 + parameter[2] * i + parameter[3])
    # y2.append(parameter[0] * i ** 2 + parameter[1] * i + parameter[2])
correlation = np.corrcoef(x, y)
p = np.poly1d(parameter)
# plt.plot(x, y2, color='g')
fig, ax = plt.subplots()
plt.scatter(x, y)
ax.plot(x, y2, color='g')
ax.xaxis.set_major_formatter(ticker.PercentFormatter(xmax=1))
plt.show()
plt.xlabel('Percentage of false satisfying MGs')
plt.ylabel('Improvement ratio of accuracy rate')
# plt.title('KNN')
plt.show()

# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[0:2], improve[0:2])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('TSQ')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[2:9], improve[2:9])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('DM')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[9:13], improve[9:13])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('SMM')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[13:18], improve[13:18])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('Tcas')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[18:21], improve[18:21])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('KNN')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[21:32], improve[21:32])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('MATH')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[32:34], improve[32:34])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('PT')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[34:37], improve[34:37])
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('Num')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[37:42], improve[37:42])
# # plt.legend(('TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA'))
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.title('DNA')
# plt.show()
#
# plt.figure()
# plt.scatter(value_fs[0:2], improve[0:2])
# plt.scatter(value_fs[2:9], improve[2:9])
# plt.scatter(value_fs[9:13], improve[9:13])
# plt.scatter(value_fs[13:18], improve[13:18])
# plt.scatter(value_fs[18:21], improve[18:21])
# plt.scatter(value_fs[21:32], improve[21:32])
# plt.scatter(value_fs[32:34], improve[32:34])
# plt.scatter(value_fs[34:37], improve[34:37])
# plt.scatter(value_fs[37:42], improve[37:42])
# plt.legend(('TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA'))
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# # plt.title('DNA')
# plt.show()

# plt.figure()
# plt.scatter(value_fs, improve)
# # plt.legend(('TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA'))
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# # plt.title('DNA')
# plt.show()

#
# Data = {}
# Datalist = []
# for k in program:
#     ws = wb[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == len(Formula):
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#         #     continue
#         data = {
#             Formula[i]: Value[i]
#         }
#         datadist.update(data)
#     Datalist.append(datadist)
#
#
# for i in range(len(Formula)):
#     # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#     #     continue
#     dd = []
#     value = []
#     for j in range(len(program)):
#         dd += Datalist[j][Formula[i]]
#     data = {
#         Formula[i]: round(np.mean(dd), 2)
#     }
#     Data.update(data)
#
# z = zip(Data.values(), Data.keys())
# z = sorted(z, reverse=True)
# values = []
# for i in range(len(z)):
#     values.append(z[i][0])
# ave = round(np.mean(values), 2)


# path = '/Applications/work/data/MT/FAILTIM/Result/Bonferroni.xlsx'
# wb = load_workbook(path)
# string = 'Sheet1'
# del wb[string]
# ws = wb.create_sheet(string)
# row = 1
# datadist = {}
# tablelist = {'Formulas': ['Mean', 'St.Dev.']}
# datadist.update(tablelist)
# mean = []
# std = []
# for i in range(len(Formula)):
#     # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#     #     continue
#     dd = []
#     value = []
#     for j in range(len(program)):
#         dd += Datalist[j][Formula[i]]
#     data = {
#         Formula[i]: dd
#     }
#     Data.update(data)
#     mean.append(round(np.mean(dd), 2))
#     std.append(round(np.std(dd), 2))
#
#     value = [round(np.mean(dd), 2), round(np.std(dd), 2)]
#     data = {
#         Formula[i]: value
#     }
#     datadist.update(data)

# for i, j in datadist.items():  # i--公式名称, j--指标值
#     ws.cell(row, 1).value = i  # 添加第 1 列的数据
#     for col in range(2, len(j) + 2):  # values列表中索引
#         ws.cell(row, col).value = j[col - 2]
#     row += 1  # 行数

# wb.save(path)
# theta
# sp2 = 0
# for i in range(len(Formula)):
#     sp2 += (57 * std[i] ** 2)
# sp2 /= (57 * 30)
# t = 3.865638
# theta = t * np.sqrt(sp2) * np.sqrt(2 / 58)

# path = '/Applications/work/data/MT/FAILTIM/Result/t test.xlsx'
# wb = load_workbook(path)
# string = 't test2'
# del wb[string]
# ws = wb.create_sheet(string)
# row = 1
# datadist = {}
# tablelist = {'formula pairs': ['t-score', 'degree of freedom', 'theta']}
# datadist.update(tablelist)
# t_score = []
# fp = []
# mm = 435
# for m in Formula:
#     for n in Formula:
#         s = 0
#         if m == n:
#             continue
#         if not fp:
#             pass
#         else:
#             for k in fp:
#                 if m in k and n in k:
#                     s = 1
#                     break
#         if s == 1:
#             continue
#         string = m + '-' + n
#         fp.append(string)
#         sp2 = 0
#         x1 = Data[m]
#         x2 = Data[n]
#         for i in range(len(x1)):
#             sp2 += ((x1[i] - np.mean(x1)) ** 2 + (x2[i] - np.mean(x2)) ** 2)
#         sp2 /= (2 * len(x1) - 2)
#         t = abs(np.mean(x1) - np.mean(x2)) / np.sqrt(2 * sp2 / len(x1))
#         t_score.append(t)
#
#         value = [round(t, 4), 57, 0.05/mm]
#         mm -= 1
#         data = {
#             string: value
#         }
#         datadist.update(data)
#
# for i, j in datadist.items():  # i--公式名称, j--指标值
#     ws.cell(row, 1).value = i  # 添加第 1 列的数据
#     for col in range(2, len(j) + 2):  # values列表中索引
#         ws.cell(row, col).value = j[col - 2]
#     row += 1  # 行数
# wb.save(path)

# from scipy import stats
# stats.ttest_ind_from_stats(91.53, 14.6, 58, 82.58, 17.49, 58, equal_var=True)

