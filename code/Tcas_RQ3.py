'''
相比V12, RQ2

Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
from Tcas import *
import random
import json
random.seed(1)
import numpy as np
from publicFun import *
from openpyxl import load_workbook

def riskIndex(argv, dynamic):
    MGS = []
    index = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = TCAS().Tcas(testcase[i])  # oracle
        result_s_m = dynamic.Tcas(testcase[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    # p or f统计完

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    # 统计全部的v和s
    sum_s = 0
    sum_v = 0
    for i in range(len(MGS)):
        sum_s += MGS[i].count(0)
        sum_v += MGS[i].count(1)
    # if not sum_s + sum_v == (len(MGS) * len(MGS[0])):
    #     print('异常')

    for i in range(len(testcase)):
        if i == 0:
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
        elif i <= len(MGS[0]):
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
            if MGS[0][i-1] == 0:
                es += 1
                ns -= 1
            elif MGS[0][i-1] == 1:
                ev += 1
                nv -= 1
        else:
            t = int((i-1) / len(MGS[0]))
            if MGS[t][(i-1) % len(MGS[0])] == 0:
                es = 1
                ns = sum_s - es
                ev = 0
                nv = sum_v
            elif MGS[t][(i-1) % len(MGS[0])] == 1:
                ev = 1
                nv = sum_v - ev
                es = 0
                ns = sum_s
            else:
                es = 0
                ns = sum_s
                ev = 0
                nv = sum_v
        index.append([ev, es, nv, ns])
    return MGS, index, Result

def selectcase(source_case_set):
    dynamic = TCASFactory("TCAS").getTCAS()
    bad_index = []
    for i in range(len(source_case_set)):
        test_case = source_case_set[i]
        MG, follow_case = MTG(test_case, dynamic)
        if 2 in MG:  # 无解的和不在考虑范围的测试用例
            bad_index.append(i)
            continue
        for j in range(len(follow_case)):
            MG, ffollow_case = MTG(follow_case[j], dynamic)
            if 2 in MG:  # 无解的和不在考虑范围的测试用例
                bad_index.append(i)
                break
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in bad_index)]
    return source_case_set

def getSource(dynamic):
    source_case_set = []
    Follow = []
    path = r"D:\程序\博士\MT\P1\tcas\testplans.alt\universe.txt"
    with open(path, 'r') as f:
        list_read = f.readlines()
    for i in range(500):
        test_case = []
        list2 = list_read[i].split()
        for j in range(len(list2)):
            test_case.append(int(list2[j]))  # 字符串转整型
        # 获取符合MR要求的源测试用例
        if test_case[6] <= 3:
            source_case_set.append(test_case)
    for i in range(len(source_case_set)):
        MGS, follow_case = MTG(source_case_set[i], dynamic)
        Follow.append(follow_case)

    for i in range(len(Follow)):
        for j in range(len(Follow[i])):
            source_case_set.append(Follow[i][j])
    source_case_set = selectcase(source_case_set)  # 去除了不满足MR的测试用例

    MG = []
    pf = []
    Index = []
    delete = []
    for i in range(len(source_case_set)):
        MGS, index, Result = riskIndex(source_case_set[i], dynamic)  # MG里面有3
        for j in range(len(index)):
            if index[j][0] + index[j][2] == 0 or index[j][1] + index[j][3] == 0:  # 去除全部为1或者0的测试用例
                delete.append(i)
        MG.append(MGS)
        pf.append(Result)
        Index.append(index)
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in delete)]
    MG = [MG[i] for i in range(len(MG)) if (i not in delete)]
    Index = [Index[i] for i in range(len(Index)) if (i not in delete)]
    pf = [pf[i] for i in range(len(pf)) if (i not in delete)]
    source_case_set, MG, pf, Index = modifycase_fairness2(source_case_set, MG, pf, Index)

    return source_case_set, MG, pf, Index

def getRisk(source_case_set):
    MG = []
    pf = []
    Index = []
    for i in range(len(source_case_set)):
        MGS, index, Result = riskIndex(source_case_set[i], dynamic)  # MG里面有3
        MG.append(MGS)
        pf.append(Result)
        Index.append(index)

    Risk = []
    for i in range(len(Index)):
        risk = []
        index = Index[i]
        for j in range(len(index)):
            try:
                if index[j][0] + index[j][1] == 0:
                    formula = [-1000 for i in range(30)]
                elif index[j][0] + index[j][2] == 0:  # v = 0
                    formula = [-1000 for i in range(30)]
                elif index[j][1] + index[j][3] == 0:  # s = 0
                    formula = [1000 for i in range(30)]
                else:
                    formula = riskformula(index[j])
                risk.append(formula)
            except:
                print(index[j])
        Risk.append(risk)
    return pf, MG, Index, Risk


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/RQ3_3.xlsx'
    wb = load_workbook(path)
    string = 'tcas'
    del wb[string]
    ws = wb.create_sheet(string)
    MGSet = []
    pfset = []
    for mu in range(1, 12):
        # dynamic = TCASFactory("Mutant" + str(mu)).getTCAS()
        # 读测试用例
        # with open('Result/Tcas/mutant' + str(mu) + '_rq1.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # source_case_set = data['source_case_set']
        #
        # pf, MG, Index, Risk = getRisk(source_case_set)
        # data = {
        #         'pf': pf, 'MG':MG, 'Risk':Risk, 'Index':Index
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('Result/Tcas/mutant' + str(mu) + '_rq3.json', 'w') as f:
        #     json.dump(json_str, f)
        # 读数据
        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']
        pfset.append(pf)
        MGSet.append(MG)

        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics_rq3_3(row, ws, mu, MG, pf)
    wb.save(path)


