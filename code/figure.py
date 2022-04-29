import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from matplotlib import ticker
from scipy.interpolate import make_interp_spline
import json

# x = ['Arithmetic Mean', 'Kulczynski2', 'AMPLE2', 'Ochiai', 'Wong3']
#
# Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice','Goodman','Tarantula',
#            'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto','Hamming etc.', 'Euclid', 'Scott',
#            'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
# path = '/Applications/work/data/MT/FAILTIM/Result/RQ1_3.xlsx'
# wb = load_workbook(path)
# program = ['trisquare', 'determinant', 'sparsemat', 'tcas', 'knn', 'MATH', 'printtokens',  'Num', 'DNA']
# Data = {}
# Datalist = []
# for k in program:
#     ws = wb[k]
#     max_row = ws.max_row
#     datadist = {}
#     Value = []
#     for i in range(len(Formula)):
#         Value.append([])
#     t = 0
#     n = 2
#     while True:
#         value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
#         Value[t].append(value)
#         t = t + 1
#         n = n + 1
#         if t == 30:
#             t = 0
#             n = n + 3
#         if n > max_row:
#             break
#     for i in range(len(Formula)):
#         # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#         #     continue
#         if not Formula[i] in x:
#             continue
#         data = {
#             Formula[i]: Value[i]
#         }
#         datadist.update(data)
#     Datalist.append(datadist)
#
# # sort = []
# # for i in range(len(Datalist)):
# #     z = zip(Datalist[i].values(), Datalist[i].keys())
# #     z = sorted(z, reverse=True)
# #     sort.append(z)
#
# for i in range(len(Formula)):
#     # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
#     #     continue
#     if not Formula[i] in x:
#         continue
#     dd = []
#     for j in range(len(program)):
#         dd += Datalist[j][Formula[i]]
#     data = {
#         Formula[i]: dd
#     }
#     Data.update(data)
#
# for i in range(len(Data)):
#     Data[x[i]] = sorted(Data[x[i]])
# formulas = ['0', 'Kulczynski2', 'Ochiai', 'AMPLE2', 'Wong3', 'Arithmetic Mean']
# ave = [0, 90.85, 90.7, 90.72, 90.65, 90.88]
# values = []
# for i in range(len(z)):
#     values.append(z[i][0])
# ave = round(np.mean(values), 2)
# #
# # data = {
# # x[0]: [1000, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900, 2500],
# # x[1]: [1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900, 2000, 2100],
# # x[2]: [1000, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900, 2000],
# # x[3]: [800, 1000, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900],
# # x[4]: [800, 1000, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900]
# # }
#
# df = pd.DataFrame(Data)
# df.plot.box()
# plt.scatter(formulas, ave, marker='*', c='r')
# plt.xlabel("Formulas")
# plt.ylabel("FI score")
# plt.grid(linestyle="--", alpha=0.3)
#
# plt.show()


# 这两行代码解决 plt 中文显示的问题
# plt.rcParams['font.sans-serif'] = ['SimHei']
# plt.rcParams['axes.unicode_minus'] = False
#
# # 输入统计数据
program = ('TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA')
# # # program = ('TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA', 'Ave')
# # hui = [95.67, 91.29, 99.15, 94.41, 84.44, 68.82, 82.58, 97.3, 87.83, 89.05]
# # best = [98.19, 93.18, 100.0, 96.87, 90.69, 81.2, 82.58, 100, 91.26, 90.88]
# # am = [98.18, 92.62, 100, 96.48, 90.35, 80.23, 71.05, 99.5, 89.52, 90.88]
# # worst = [95.63, 78.34, 97.63, 92.1, 83.37, 61.32, 58.96, 91.15, 77.82, 83.62]
# #
# # bar_width = 0.2  # 条形宽度
# # index_hui = np.arange(len(program))  # 男生条形图的横坐标
# # index_best = index_hui + bar_width  # 女生条形图的横坐标
# # index_am = index_best + bar_width # 男生条形图的横坐标
# # index_worst = index_am + bar_width  # 女生条形图的横坐标
# #
# #
# # # 使用两次 bar 函数画出两组条形图
# # plt.bar(index_hui, height=hui, width=bar_width, label='Hui’s method ')
# # plt.bar(index_best, height=best, width=bar_width, label='Best formula')
# # plt.bar(index_am, height=am, width=bar_width, label='Arithmetic Mean')
# # plt.bar(index_worst, height=worst, width=bar_width, label='Worst formula')
# #
# # plt.legend(loc='upper right')  # 显示图例
# # plt.xticks(index_hui + bar_width, program)  # 让横坐标轴刻度显示 waters 里的饮用水， index_male + bar_width/2 为横坐标轴刻度的位置
# # plt.ylabel('FI score')  # 纵坐标轴标题
# # plt.xlabel("Programs")
# # plt.ylim([50, 120])
# # # plt.title('购买饮用水情况的调查结果') # 图形标题
# # plt.show()
# all
# With = [96.5, 91.73, 99.76, 96.09, 88.2, 76.67, 69.15, 98.91, 87.85]
# Without = [96.77, 95.8, 99.78, 97.56, 97.07, 89.89, 88.39, 99.15, 96.9]
# improve = [Without[i] - With[i] for i in range(0, len(With))]
# ave = np.mean(improve)
# fs = [13.10666667, 48.12714286, 83.64142857, 22.00636364, 38.29666667, 51.61, 50.07, 12.16666667, 57.11375]
# # plt.scatter(fs, improve)
# plt.scatter(fs[0], improve[0])
# plt.scatter(fs[1], improve[1])
# plt.scatter(fs[2], improve[2])
# plt.scatter(fs[3], improve[3])
# plt.scatter(fs[4], improve[4])
# plt.scatter(fs[5], improve[5])
# plt.scatter(fs[6], improve[6])
# plt.scatter(fs[7], improve[7])
# plt.scatter(fs[8], improve[8])
# plt.legend(('TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA'))
# plt.xlabel('Ratio of false satisfying MGs')
# plt.ylabel('Improvement of accuracy ratio')
# plt.show()

# all
With = [96.5, 91.73, 99.76, 96.09, 88.2, 76.67, 69.15, 98.91, 87.85]
# Without = [96.77, 95.8, 99.78, 97.56, 97.07, 89.89, 88.39, 99.15, 96.9]
Without = [96.77, 95.8, 99.78, 97.56, 97.07, 89.89, 94.05, 99.15, 96.9]
improve = [Without[i] - With[i] for i in range(0, len(With))]
ave = np.mean(improve)
# remove bin and naish
# With = [96.56, 92.67, 99.9, 96.38, 88.4, 77.75, 68.68, 99.43, 88.56]
# Without = [96.85, 97.03, 99.92, 97.95, 97.9, 91.85, 89.08, 99.68, 98.25]
# # am
# With = [98.18, 92.62, 100.0, 96.48, 90.35, 80.23, 71.05, 99.5, 89.52]
# Without = [98.86, 99.72, 100.0, 99.98, 100.0, 99.7, 89.9, 99.75, 99.54]

bar_width = 0.3  # 条形宽度
index_With = np.arange(len(program))  # 男生条形图的横坐标
index_Without = index_With + bar_width  # 女生条形图的横坐标

# 使用两次 bar 函数画出两组条形图
# plt.bar(index_With, height=With, width=bar_width, label='with false satisfying MGs')
# plt.bar(index_Without, height=Without, width=bar_width, label='without false satisfying MGs')
# plt.legend(loc='upper right')  # 显示图例
fig, ax = plt.subplots()
ax.bar(index_With, height=With, width=bar_width, label='with false satisfying MGs')
ax.bar(index_Without, height=Without, width=bar_width, label='without false satisfying MGs')
plt.legend(loc='upper right')  # 显示图例
ax.yaxis.set_major_formatter(ticker.PercentFormatter())
plt.xticks([index - 0.15 for index in (index_With + bar_width)], program)  # 让横坐标轴刻度显示 waters 里的饮用水， index_male + bar_width/2 为横坐标轴刻度的位置
plt.ylabel('Accurary rate')  # 纵坐标轴标题
plt.xlabel("Programs")
plt.ylim([60, 110])
# plt.title('购买饮用水情况的调查结果') # 图形标题
plt.show()


# x_data1 = ['d≥1', 'd≥2', 'd≥3', 'd≥4', 'd≥5', 'd≥6', 'd≥7', 'd≥8', 'd≥9', 'd≥10', 'd≥11']
# x_data2 = ['d≥1', 'd≥2', 'd≥3', 'd≥4', 'd≥5', 'd≥6', 'd≥7', 'd≥8', 'd≥9']
# x_data3 = ['d≥1', 'd≥2', 'd≥3', 'd≥4', 'd≥5', 'd≥6']
# x_data4 = ['d≥1', 'd≥2', 'd≥3', 'd≥4', 'd≥5', 'd≥6', 'd≥7', 'd≥8', 'd≥9', 'd≥10']
# # all
# Trisuqare = [96.77,	96.78, 96.91, 97.09, 97.39, 97.75, 98.24, 98.87, 99.33]
# Determinant = [95.8, 93.44,	94.16, 93.82, 95.13, 98.1, 98.28, 98.13, 98.83,	96.67]
# Sparsemat = [99.78,	99.85,	99.87, 99.86, 99.88, 99.88, 99.39, 100.0, 100.0]
# Tcas = [97.56, 99.38, 99.53, 99.64,	99.67, 99.55, 99.66, 99.65, 100.0]
# KNN = [97.07, 94.68, 97.19,	97.97, 98.33, 96.67]
# MATH = [89.89, 91.32, 92.26, 93.71,	94.96, 95.82, 96.61, 97.19,	97.14, 97.61, 97.68]
# Print_tokens = [88.39,	90.81,	92.0, 93.62, 94.83,	95.42, 93.18, 93.37, 91.06]
# Number = [99.15, 99.19,	99.19, 99.02, 98.78, 98.7]
# DNA = [96.9, 96.67,	97.15, 97.36, 97.71, 98.21,	98.26, 98.34, 97.23, 96.96,	96.67]
#
# # create integers from strings
# # idx1 = range(len(x_data1))
# # xnew1 = np.linspace(min(idx1), max(idx1), 300)
# # idx2 = range(len(x_data2))
# # xnew2 = np.linspace(min(idx2), max(idx2), 300)
# # idx3 = range(len(x_data3))
# # xnew3 = np.linspace(min(idx3), max(idx3), 300)
# # idx4 = range(len(x_data4))
# # xnew4 = np.linspace(min(idx4), max(idx4), 300)
# # # interpolation
# # spltr = make_interp_spline(idx2, Trisuqare, k=3)
# # smoothtr = spltr(xnew2)
# # spldm = make_interp_spline(idx4, Determinant, k=3)
# # smoothdm = spldm(xnew4)
# # splsmm = make_interp_spline(idx2, Sparsemat, k=3)
# # smoothsmm = splsmm(xnew2)
# # spltc = make_interp_spline(idx2, Tcas, k=3)
# # smoothtc = spltc(xnew2)
# # splk = make_interp_spline(idx3, KNN, k=3)
# # smoothk = splk(xnew3)
# # splm = make_interp_spline(idx1, MATH, k=3)
# # smoothm = splm(xnew1)
# # splp = make_interp_spline(idx2, Print_tokens, k=3)
# # smoothp = splp(xnew2)
# # spln = make_interp_spline(idx3, Number, k=3)
# # smoothn = spln(xnew3)
# # spldna = make_interp_spline(idx1, DNA, k=3)
# # smoothdna = spldna(xnew1)
#
# # am
# # Trisuqare = [98.86,	98.87,	98.91,	98.93,	98.99,	99.1,	99.28,	99.34,	100.0]
# # Determinant = [99.72,	99.6,	99.79,	99.87,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # Sparsemat = [100.0,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # Tcas = [99.98,	99.98,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # KNN = [100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # MATH = [99.7,	99.69,	99.84,	99.93,	99.98,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # Print_tokens = [89.9,	91.35,	93.6,	94.94,	95.03,	96.34,	94.1,	94.79,	93.18]
# # Number = [99.75,	99.7,	100.0,	100.0,	100.0,	100.0]
# # DNA = [99.54,	99.47,	99.65,	99.87,	99.92,	99.85,	100.0,	100.0,	100.0,	100.0,	100.0]
# # remove bin and n1
# # Trisuqare = [96.85,	96.86,	97.02,	97.24,	97.57,	97.96,	98.53,	99.3,	100.0]
# # Determinant = [97.03,	94.7,	95.56,	95.36,	96.86,	99.88,	99.98,	100.0,	100.0,	100.0]
# # Sparsemat = [99.92,	99.98,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # Tcas = [97.95,	99.72,	99.88,	100.0,	100.0,	100.0,	100.0,	100.0,	100.0]
# # KNN = [97.9,	95.51,	98.98,	99.76,	100.0,	100.0]
# # MATH = [91.85,	93.08,	94.1,	95.68,	97.06,	98.24,	99.25,	99.81,	99.99,	100.0,	100.0]
# # Print_tokens = [89.08,	91.43,	92.72,	94.26,	95.66,	96.27,	94.01,	94.79,	93.18]
# # Number = [99.68,	99.73,	99.82,	99.95,	100.0,	100.0]
# # DNA = [98.25,	98.18,	98.75,	99.02,	99.49,	99.72,	100.0,	100.0,	100.0,	100.0,	100.0]
#
# plt.plot(x_data2, Trisuqare, linestyle='-')
# plt.plot(x_data4, Determinant, linestyle='-')
# plt.plot(x_data2, Sparsemat, linestyle='-')
# plt.plot(x_data2, Tcas, linestyle='-')
# plt.plot(x_data3, KNN, linestyle='-')
# plt.plot(x_data1, MATH, linestyle='-')
# plt.plot(x_data2, Print_tokens, linestyle='-')
# plt.plot(x_data3, Number, linestyle='-')
# plt.plot(x_data1, DNA, linestyle='-')
# # plt.plot(xnew2, smoothtr)
# # plt.plot(xnew4, smoothdm)
# # plt.plot(xnew2, smoothsmm)
# # plt.plot(xnew2, smoothtc)
# # plt.plot(xnew3, smoothk)
# # plt.plot(xnew1, smoothm)
# # plt.plot(xnew2, smoothp)
# # plt.plot(xnew3, smoothn)
# # plt.plot(xnew1, smoothdna)
# # plt.xticks(idx1, x_data1)
# plt.ylabel('FI score')  # 纵坐标轴标题
# plt.xlabel("Data sufficiency")
# # plt.ylim([80, 100])
# # my_font = fm.FontProperties(fname="/usr/share/fonts/wqy-microhei/wqy-microhei.ttc")
# plt.legend(labels=['TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT', 'Num', 'DNA'])
#
# plt.show()

