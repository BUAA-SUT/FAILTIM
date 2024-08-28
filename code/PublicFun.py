import numpy as np
from openpyxl import load_workbook


def riskformula(index, t):
    ev = index[0]
    es = index[1]
    nv = index[2]
    ns = index[3]
    F = ev + nv
    P = es + ns
    if t == 0:
        if ev < F:
            formula = -1
        elif ev == F and F != 0:
            formula = ns
        else:
            formula = -1
    elif t == 1:
        formula = ev - es / (es + ns + 1)
    elif t == 2:
        formula = ev
    elif t == 3:
        formula = ev/(ev + es + nv + ns)
    elif t == 4:
        if ev < F:
            formula = 0
        elif ev == F and F != 0:
            formula = 1
        else:
            formula = 0
    elif t == 5:
        formula = ev / (ev + es + nv)
    elif t == 6:
        formula = ev / (ev + 2 * (nv + es))
    elif t == 7:
        formula = 2 * ev / (2 * ev + nv + es)
    elif t == 8:
        formula = 2 * ev / (ev + nv + es)
    elif t == 9:
        formula = (2 * ev - nv - es) / (2 * ev + nv + es)
    elif t == 10:
        formula = (ev / (ev + nv)) / (ev / (ev + nv) + es / (es + ns))
    elif t == 11:
        formula = ev / (ev + es)
    elif t == 12:
        formula = (ev / (ev + es)) - ((ev + nv) / (ev + nv + es + ns))
    elif t == 13:
        formula = ev - es
    elif t == 14:
        formula = (ev + ns - nv - es) / (ev + nv + es + ns)
    elif t == 15:
        formula = (ev + ns) / (ev + nv + es + ns)
    elif t == 16:
        formula = 2 * (ev + ns) / (2 * (ev + ns) + nv + es)
    elif t == 17:
        formula = (ev + ns) / (ev + ns + 2 * (nv + es))
    elif t == 18:
        formula = ev + ns
    elif t == 19:
        formula = np.sqrt(ev + ns)
    elif t == 20:
        formula = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / ((2 * ev + nv + es) * (2 * ns + nv + es))
    elif t == 21:
        formula = 0.5 * (ev / (2 * ev + nv + es) + ns / (2 * ns + nv + es))
    elif t == 22:
        formula = 0.5 * (ev / (ev + nv) + ev / (ev + es))
    elif t == 23:
        formula = ev / ((ev + nv) * (ev + es)) ** 0.5
    elif t == 24:
        formula = ev / (ev + ns + 2 * (nv + es))
    elif t == 25:
        formula = ev / (ev + nv) - es / (es + ns)
    elif t == 26:
        if es <= 2:
            formula = ev - es
        elif 2 < es <= 10:
            formula = ev - 2 - 0.1 * (es - 2)
        else:
            formula = ev - 2.8 - 0.001 * (es - 10)
    elif t == 27:
        formula = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + nv) + (ev + nv) * (es + ns))
    elif t == 28:
        formula = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + es) + (ev + nv) * (nv + ns))
    else:
        formula = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / (2 * ev + nv + es + 2 * ns + nv + es)

    return formula


def getSus(index, t):
    formula = 0
    try:
        formula = riskformula(index, t)
    except:
        if index[0] + index[1] == 0:  # e = 0
            formula = -1000
        elif index[0] + index[2] == 0:  # v = 0
            formula = -1000
        elif index[1] + index[3] == 0:  # s = 0
            formula = 1000
    return formula


def modifycase_fairness(source_case_set, MG):
    # result全为0或者全为1就删掉对应的测试用例
    index = []
    for i in range(len(MG)):
        if MG[i][0].count(0) + MG[i][0].count(3) == len(MG[i][0]) or MG[i][0].count(1) == len(MG[i][0]):
            index.append(i)
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in index)]
    return source_case_set


def modifycase_fairness2(MG, pf):
    # result全为0或者全为1就删掉对应的测试用例
    index = []
    for i in range(len(MG)):
        if MG[i][0].count(0) + MG[i][0].count(3) == len(MG[i][0]) or MG[i][0].count(1) == len(MG[i][0]):
            index.append(i)
    MG = [MG[i] for i in range(len(MG)) if (i not in index)]
    pf = [pf[i] for i in range(len(pf)) if (i not in index)]
    return MG, pf


def getMetrics(row, ws, mu, MG, pf):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'Ratio1', 'Ratio2', 'Ratio3',
                                      'VMG(%)', 'failed test case(%)', 'false satisfied MG(%)', 'fs MG / all(%)', 'fs new']}
    datadist.update(tablelist)
    for t in range(len(Formula)):
        V_MGlist = []
        FFSlist = []
        FFSalist = []
        FFSnewlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        r1list = []
        r2list = []
        r3list = []
        Failed = []
        for k in range(len(MG)):  # k代表source的个数
            r1 = []
            r2 = []
            r3 = []
            V_MG = 0
            V_MG1 = 0
            S_MG1 = 0
            ffsnew = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG
            percent_FFSa = FFS / Total_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            FFSalist.append(percent_FFSa)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                flag = 0
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        # case = 0
                        V_MG1 += 1

                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        if t == 30:
                            sus_a = nv_b
                            sus_b = nv_a
                        else:
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)

                        cta = ev_a / (sum_s + sum_v)
                        ctb = ev_b / (sum_s + sum_v)
                        ctab = (ev_a + ev_b - 1) / (sum_s + sum_v)
                        r1.append(cta)
                        r2.append(ctb)
                        r3.append(ctab)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_2 += 1
                            else:
                                p3_2 += 1
                        elif pf[k][i]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1
                            else:
                                p3_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1
                        else:
                            if sus_a < sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1
                            else:
                                p3_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4
            if S_MG1 == 0:
                percent_ffsnew = 0
            else:
                percent_ffsnew = ffsnew / S_MG1

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)
            r1list.append(np.mean(r1))
            r2list.append(np.mean(r2))
            r3list.append(np.mean(r3))
            FFSnewlist.append(np.mean(percent_ffsnew))


        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2), round(np.mean(r1list) * 100, 2),
                 round(np.mean(r2list) * 100, 2), round(np.mean(r3list) * 100, 2), round(np.mean(V_MGlist) * 100, 2), round(np.mean(Failed) * 100, 2),
                 round(np.mean(FFSlist) * 100, 2), round(np.mean(FFSalist) * 100, 2), round(np.mean(FFSnewlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_nofs(row, ws, mu, MG, pf):
    """
    去除fs
    """

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0)
                        sum_v = MG[k][i].count(1)
                        sum_s += MG[k][j+1].count(0)
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                            case = 4
                        elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                            if pf[k][i] and sus_a > sus_b:
                                Fp += 1
                                case = 1
                                # print(k, i, j, index_a, index_b, sus_a, sus_b)
                            elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                                Fp += 1
                                case = 1
                                # print(k, i, j, index_a, index_b, sus_a, sus_b)
                            else:
                                case = 3
                                Pf += 1  # p比f高

                        else:  # 只有一个failed, 但是相等
                            FP += 1
                            case = 2
                            # print(sus_a[t], sus_b[t])
                            # print(index_a, index_b)
                            # print(k, i, j)
                        cases.append(case)

            if len(cases) == 0:
                continue
            index.append(cases)

            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)
        Index.append(index)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row
