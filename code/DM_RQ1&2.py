'''
RQ1
Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
import numpy as np
import random
import os
import json
random.seed(1)
from DM import *
from publicFun import *
from openpyxl import load_workbook

def riskIndex(argv, dynamic):
    MGS = []
    # index = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = DeterMinant().Determinant(testcase[i], n)  # oracle
        result_s_m = dynamic.Determinant(testcase[i], n)
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    # p or f统计完
    # 随机去除一些MG
    for i in range(1, len(MGS)):  # 第一组不变
        t = random.randint(1, len(MGS[i])-1)  # 去几个
        a = [i for i in range(len(MGS[i]))]
        random.shuffle(a)
        b = a[:t]
        for j in b:
            MGS[i][j] = 4

    return MGS, Result


def getSource(dynamic):
    Follow = []
    source_case_set = [[2, 7, 8, 1, 2, 7, 8, 8, 1], [7, 4, 6, 8, 4, 0, 5, 0, 0], [0, 4, 6, 8, 0, 0, 5, 0, 0],
                       [2, 0, 1, 3, 8, 2, 6, 4, 9], [2, 0, 0, 1, 5, 0, 2, 4, 6], [0, 0, 0, 1, 0, 0, 2, 4, 0]]   #分支312312

    for i in range(len(source_case_set)):
        MGS, follow_case = MTG(source_case_set[i], dynamic)
        Follow.append(follow_case)

    for i in range(len(Follow)):
        for j in range(len(Follow[i])):
            source_case_set.append(Follow[i][j])
    MG = []
    for i in range(len(source_case_set)):
        MGS, Result = riskIndex(source_case_set[i], dynamic)
        MG.append(MGS)
    source_case_set = modifycase_fairness(source_case_set, MG)

    if len(source_case_set) == 0:
        print('测试用例集不满足要求')

    return source_case_set


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/correlation_with3.xlsx'
    wb = load_workbook(path)
    string = 'determinant'
    del wb[string]
    ws = wb.create_sheet(string)
    n = 3
    source = []
    for mu in range(1, 8):
        # dynamic = DeterFactory("Mutant" + str(mu)).getDeter()
        # source_case_set = getSource(dynamic)
        # # data = {
        # #         'source_case_set': source_case_set
        # # }
        # # # 将数据存下来
        # # json_str = json.dumps(data)
        # # with open('Result/Determinant/mutant' + str(mu) + '_rq1.json', 'w') as f:
        # #     json.dump(json_str, f)
        # # pf, MG, Risk = getRisk(source_case_set)
        #
        #
        # # 读数据
        # with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) +  '_rq1.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # source_case_set = data['source_case_set']
        # source.append(source_case_set)
        # MG_set = []
        # pf_set = []
        # for i in range(len(source_case_set)):
        #     MG, pf = riskIndex(source_case_set[i], dynamic)
        #     MG_set.append(MG)
        #     pf_set.append(pf)
        # data = {
        #         'source_case_set': source_case_set, 'pf': pf_set, 'MG': MG_set
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('Result/Determinant/mutant' + str(mu) + '_rq1.json', 'w') as f:
        #     json.dump(json_str, f)
        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']

        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics_rq1_3_new(row, ws, mu, MG, pf)

        # data = {
        #     'cases': cases
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open(
        #         '/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_newcases.json',
        #         'w') as f:
        #     json.dump(json_str, f)

    wb.save(path)
