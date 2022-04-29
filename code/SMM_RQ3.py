'''
相比V7, RQ2
Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
import SMM as smm
import sys
import math
import numpy as np
import json
import scipy.sparse as ss
from publicFun import *
import random
from openpyxl import load_workbook
random.seed(1)

def mat_transpose(A):
    row = len(A)
    column = len(A[0])
    B = []
    for i in range(column):
        temp = []
        for j in range(0, row):
            temp.append(A[j][i])
        B.append(temp)
    return B


def mat_copy(A):
    B = []
    for i in range(len(A)):
        temp = []
        for j in range(len(A[0])):
            temp.append(A[i][j])
        B.append(temp)
    return B


def AssertMRViolation(ftc_output, ftc_expected_output):
    d = 0.00001
    row = len(ftc_expected_output)
    col = len(ftc_expected_output[0])
    if not (row == len(ftc_output) and col == len(ftc_output[0])):
        return True
    for i in range(row):
        for j in range(col):
            if math.fabs(ftc_output[i][j] - ftc_expected_output[i][j]) >= d:
                return True
    return False


def getIdentityMatrix(dim):
    I = []
    for i in range(dim):
        temp = []
        for j in range(dim):
            if i == j:
                temp.append(1)
            else:
                temp.append(0)
        I.append(temp)
    return I


"""Return a square matrix, which is transformed by exchanging two rows of an Identity matrix."""


def getPMatrix(dim):
    if dim < 2:
        print("Cannot compound, since matrix has less than two rows!")
        sys.exit(-1)
    P = getIdentityMatrix(dim)
    temp = P[0]
    P[0] = P[1]
    P[1] = temp
    return P


"""Return a square Q matrix, whose principle diagonal elements are all constant c."""


def getQMatrix(dim, constant):
    Q = []
    for i in range(dim):
        temp = []
        for j in range(dim):
            if i == j:
                temp.append(constant)
            else:
                temp.append(0)
        Q.append(temp)
    return Q


"""Return a matrix, whose elements are all scaled with the given scalar. """


def mat_multiple_constant(matrix, scalar):
    scaled_matrix = mat_copy(matrix)
    for i in range(len(scaled_matrix)):
        for j in range(len(scaled_matrix[0])):
            scaled_matrix[i][j] = scaled_matrix[i][j] * scalar
    return scaled_matrix


def mat_addition(A, B):
    C = []
    if not (len(A) == len(B) and len(A[0]) == len(B[0])):
        print("Can not compound!\n Matrix addition failure, since they do not have the same dimension.")
        sys.exit(-1)
    for i in range(len(A)):
        temp = []
        for j in range(len(A[0])):
            temp.append(A[i][j] + B[i][j])
        C.append(temp)
    return C


def MR1(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    ftc_A = mat_transpose(otc_B)
    ftc_B = mat_transpose(otc_A)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_transpose(ori_output)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR2(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    P = getPMatrix(len(otc_A))
    ftc_A, symbol = smm.MatMul((P, otc_A))
    ftc_B = mat_copy(otc_B)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((P, ori_output))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR3(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    P = getPMatrix(len(otc_B[0]))
    ftc_A = mat_copy(otc_A)
    ftc_B, symbol = smm.MatMul((otc_B, P))
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((ori_output, P))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR4(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    Q = getQMatrix(len(otc_A), 3)
    ftc_A, symbol = smm.MatMul((Q, otc_A))
    ftc_B = mat_copy(otc_B)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((Q, ori_output))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR5(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    Q = getQMatrix(len(otc_A), 4)
    ftc_B, symbol = smm.MatMul((otc_B, Q))
    ftc_A = mat_copy(otc_A)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((ori_output, Q))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR6(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    scalar = 6
    ftc_A = mat_multiple_constant(otc_A, scalar)
    ftc_B = mat_copy(otc_B)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_multiple_constant(ori_output, scalar)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR7(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    scalar = 7
    ftc_B = mat_multiple_constant(otc_B, scalar)
    ftc_A = mat_copy(otc_A)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_multiple_constant(ori_output, scalar)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR8(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    dim = len(otc_A)
    I = getIdentityMatrix(dim)
    ftc_A = mat_addition(otc_A, I)
    ftc_B = mat_copy(otc_B)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_addition(ori_output, otc_B)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR9(mat, dynamic):
    otc_A = mat[0]
    otc_B = mat[1]
    dim = len(otc_B)
    I = getIdentityMatrix(dim)
    ftc_A = mat_copy(otc_A)
    ftc_B = mat_addition(otc_B, I)
    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_addition(otc_A, ori_output)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR10(mat, dynamic):
    # 先MR1
    otc_A = mat[0]
    otc_B = mat[1]
    ftc_A = mat_transpose(otc_B)
    ftc_B = mat_transpose(otc_A)
    # ori_output, symbol = matmultiple.MatMul((otc_A, otc_B))
    # mrs_output, symbol = matmultiple.MatMul((ftc_A, ftc_B))
    # ftc_expected_output = mat_transpose(ori_output)    # 再MR2
    otc_A = ftc_A
    otc_B = ftc_B
    P = getPMatrix(len(otc_A))
    ftc_A, symbol = smm.MatMul((P, otc_A))
    ftc_B = mat_copy(otc_B)

    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((P, ori_output))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR11(mat, dynamic):
    # 先MR1
    otc_A = mat[0]
    otc_B = mat[1]
    ftc_A = mat_transpose(otc_B)
    ftc_B = mat_transpose(otc_A)
    # ori_output, symbol = matmultiple.MatMul((otc_A, otc_B))
    # mrs_output, symbol = matmultiple.MatMul((ftc_A, ftc_B))
    # ftc_expected_output = mat_transpose(ori_output)

    # 再MR3
    otc_A = ftc_A
    otc_B = ftc_B
    P = getPMatrix(len(otc_B[0]))
    ftc_A = mat_copy(otc_A)
    ftc_B, symbol = smm.MatMul((otc_B, P))

    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((ori_output, P))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR12(mat, dynamic):
    # 先MR2
    otc_A = mat[0]
    otc_B = mat[1]
    P = getPMatrix(len(otc_A))
    ftc_A, symbol = smm.MatMul((P, otc_A))
    ftc_B = mat_copy(otc_B)

    # 再MR1
    otc_A = ftc_A
    otc_B = ftc_B
    ftc_A = mat_transpose(otc_B)
    ftc_B = mat_transpose(otc_A)

    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_transpose(ori_output)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR13(mat, dynamic):
    # 先MR2
    otc_A = mat[0]
    otc_B = mat[1]
    P = getPMatrix(len(otc_A))
    ftc_A, symbol = smm.MatMul((P, otc_A))
    ftc_B = mat_copy(otc_B)

    # 再MR3
    otc_A = ftc_A
    otc_B = ftc_B
    P = getPMatrix(len(otc_B[0]))
    ftc_A = mat_copy(otc_A)
    ftc_B, symbol = smm.MatMul((otc_B, P))

    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((ori_output, P))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR14(mat, dynamic):
    # 先MR3
    otc_A = mat[0]
    otc_B = mat[1]
    P = getPMatrix(len(otc_B[0]))
    ftc_A = mat_copy(otc_A)
    ftc_B, symbol = smm.MatMul((otc_B, P))

    # 再MR1
    otc_A = ftc_A
    otc_B = ftc_B
    ftc_A = mat_transpose(otc_B)
    ftc_B = mat_transpose(otc_A)

    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output = mat_transpose(ori_output)
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MR15(mat, dynamic):
    # 先MR3
    otc_A = mat[0]
    otc_B = mat[1]
    P = getPMatrix(len(otc_B[0]))
    ftc_A = mat_copy(otc_A)
    ftc_B, symbol = smm.MatMul((otc_B, P))

    # 再MR2
    otc_A = ftc_A
    otc_B = ftc_B
    P = getPMatrix(len(otc_A))
    ftc_A, symbol = smm.MatMul((P, otc_A))
    ftc_B = mat_copy(otc_B)

    ori_output, symbol = dynamic.MatMul((otc_A, otc_B))
    mrs_output, symbol = dynamic.MatMul((ftc_A, ftc_B))
    ftc_expected_output, symbol = smm.MatMul((P, ori_output))
    violation = AssertMRViolation(ftc_expected_output, mrs_output)
    if violation:  # 违反了
        return 1, (ftc_A, ftc_B)
    else:
        return 0, (ftc_A, ftc_B)


def MTG(argv, dynamic):
    source = argv
    follow_case = []
    MG = []
    current_module = sys.modules[__name__]
    for i in range(1, 16):  # MR
        result, follow = getattr(current_module, 'MR'+str(i))(source, dynamic)
        MG.append(result)
        follow_case.append(follow)
    return MG, follow_case

def riskIndex(argv, dynamic):
    MGS = []
    index = []
    Result = []
    testcase = []
    source_case = argv
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = smm.MatMul(testcase[i])  # oracle
        result_s_m = dynamic.MatMul(testcase[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    # p or f统计完

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, index, Result


def getSource(dynamic):
    source_case_set = []
    Follow = []
    for _ in range(50):  # 4个矩阵, 注意这样写不同mutant是不同数据集
        mar = []
        for _ in range(2):
            n = 4
            m = 4
            density = random.choice([0.3, 0.4, 0.5])
            matrixformat = 'coo'
            s_mar = ss.rand(m, n, density=density, format=matrixformat, dtype=None)
            s_mar_dense = s_mar.todense()
            s_mar_list = s_mar_dense.getA().tolist()
            for i in range(n):
                for j in range(m):
                    s_mar_list[i][j] = int(s_mar_list[i][j] * 10)
            mar.append(s_mar_list)
        source_case_set.append((mar[0], mar[1]))
    # 增加一些0矩阵和元素为1的矩阵
    for i in range(10):
        mar = []
        mar.append([[0] * 4 for k in range(4)])
        mar.append(source_case_set[i][0])  # 随意的
        source_case_set.append((mar[0], mar[1]))

    for i in range(10):
        mar = []
        matrix = [[0] * 4 for k in range(4)]
        for m in range(len(matrix)):
            for n in range(len(matrix[0])):
                a = random.randint(1, 10)
                if a % 2 == 0:
                    matrix[m][n] = 1
        mar.append(matrix)
        mar.append(source_case_set[i][1])  # 随意的
        source_case_set.append((mar[0], mar[1]))

    # for i in range(len(source_case_set)):
    #     MGS, follow_case = MTG(source_case_set[i], dynamic)
    #     Follow.append(follow_case)
    #
    # for i in range(len(Follow)):
    #     for j in range(len(Follow[i])):
    #         source_case_set.append(Follow[i][j])
    MG = []
    pf = []
    Index = []
    delete = []
    for i in range(len(source_case_set)):
        MGS, index, Result = riskIndex(source_case_set[i], dynamic)  # MG里面有3
        for j in range(len(index)):
            if index[j][0] + index[j][2] == 0 or index[j][1] + index[j][3] == 0:  # 去除全部为1或者0的测试用例
                delete.append(i)
        MG.append(MGS)
        pf.append(Result)
        Index.append(index)
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in delete)]
    MG = [MG[i] for i in range(len(MG)) if (i not in delete)]
    Index = [Index[i] for i in range(len(Index)) if (i not in delete)]
    pf = [pf[i] for i in range(len(pf)) if (i not in delete)]
    source_case_set, MG, pf, Index = modifycase_fairness2(source_case_set, MG, pf, Index)

    return source_case_set, MG, pf, Index

def getRisk(source_case_set):
    MG = []
    pf = []
    Index = []
    for i in range(len(source_case_set)):
        MGS, index, Result = riskIndex(source_case_set[i], dynamic)  # MG里面有3
        MG.append(MGS)
        pf.append(Result)
        Index.append(index)

    Risk = []
    for i in range(len(Index)):
        risk = []
        index = Index[i]
        for j in range(len(index)):
            try:
                if index[j][0] + index[j][1] == 0:
                    formula = [-1000 for i in range(30)]
                elif index[j][0] + index[j][2] == 0:  # v = 0
                    formula = [-1000 for i in range(30)]
                elif index[j][1] + index[j][3] == 0:  # s = 0
                    formula = [1000 for i in range(30)]
                else:
                    formula = riskformula(index[j])
                risk.append(formula)
            except:
                print(index[j])
        Risk.append(risk)
    return pf, MG, Index, Risk


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/RQ3_3.xlsx'
    wb = load_workbook(path)
    string = 'sparsemat'
    del wb[string]
    ws = wb.create_sheet(string)
    MGSet = []
    pfset = []
    for mu in range(1, 8):
        # mutant_index = "MU_" + str(mu) + "_SparseMatMul"
        # dynamic = smm.MatrixMultiple(mutant_index)
        # 读测试用例
        # with open('Result/Sparsemat/mutant' + str(mu) + '_rq1.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # source_case_set = data['source_case_set']
        #
        # pf, MG, Index, Risk = getRisk(source_case_set)
        # data = {
        #         'pf': pf, 'MG':MG, 'Risk':Risk, 'Index':Index
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('Result/Sparsemat/mutant' + str(mu) + '_rq3.json', 'w') as f:
        #     json.dump(json_str, f)
        # 读数据
        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']
        pfset.append(pf)
        MGSet.append(MG)
        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics_rq3_3(row, ws, mu, MG, pf)
    wb.save(path)
