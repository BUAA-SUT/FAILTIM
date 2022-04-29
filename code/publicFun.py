import numpy as np
from openpyxl import load_workbook
def riskformula(index): # 多个公式
    ev = index[0]
    es = index[1]
    nv = index[2]
    ns = index[3]
    F = ev + nv
    P = es + ns
    if ev < F:
        N1 = -1
        Bin = 0
    elif ev == F and F != 0:
        N1 = ns
        Bin = 1
    else:
        N1 = -1
        Bin = 0
    Jaccard = ev / (ev + es + nv)
    Aberg = ev / (ev + 2 * (nv + es))
    SDice = 2 * ev / (2 * ev + nv + es)
    Dice = 2 * ev / (ev + nv + es)
    Goodman = (2 * ev - nv - es) / (2 * ev + nv + es)
    Tar = (ev / (ev + nv)) / (ev / (ev + nv) + es / (es + ns))
    Qe = ev / (ev + es)
    CBI = (ev / (ev + es)) - ((ev + nv) / (ev + nv + es + ns))
    W2 = ev - es
    Hamann = (ev + ns - nv - es) / (ev + nv + es + ns)
    SM = (ev + ns) / (ev + nv + es + ns)
    Sokal = 2 * (ev + ns) / (2 * (ev + ns) + nv + es)
    RT = (ev + ns) / (ev + ns + 2 * (nv + es))
    Hamming = ev + ns
    Euclid = np.sqrt(ev + ns)
    Scott = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / ((2 * ev + nv + es) * (2 * ns + nv + es))
    Rogot1 = 0.5 * (ev / (2 * ev + nv + es) + ns / (2 * ns + nv + es))
    Kul2 = 0.5 * (ev / (ev + nv) + ev / (ev + es))
    Ochiai = ev / ((ev + nv) * (ev + es)) ** 0.5
    M2 = ev / (ev + ns + 2 * (nv + es))
    AMPLE2 = ev / (ev + nv) - es / (es + ns)
    if es <= 2:
        W3 = ev - es
    elif es > 2 and es <= 10:
        W3 = ev - 2 - 0.1 * (es - 2)
    else:
        W3 = ev - 2.8 - 0.001 * (es - 10)
    AM = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + nv) + (ev + nv) * (es + ns))
    Cohen = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + es) + (ev + nv) * (nv + ns))
    Fle = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / (2 * ev + nv + es + 2 * ns + nv + es)

    N2 = ev - es/(es + ns + 1)
    W1 = ev
    RR = ev/(ev + es + nv + ns)
    formula = [N1, N2, W1, RR, Bin, Jaccard, Aberg, SDice, Dice, Goodman, Tar, Qe, CBI, W2, Hamann,
               SM, Sokal, RT, Hamming, Euclid, Scott, Rogot1, Kul2, Ochiai, M2, AMPLE2, W3, AM, Cohen, Fle]
    return formula


def riskformula_new(index, t):  # 多个公式
    ev = index[0]
    es = index[1]
    nv = index[2]
    ns = index[3]
    ev_only = nv
    F = ev + nv
    P = es + ns
    if t == 0:
        if ev < F:
            formula = -1
        elif ev == F and F != 0:
            formula = ns
        else:
            formula = -1
    elif t == 1:
        formula = ev - es / (es + ns + 1)
    elif t == 2:
        formula = ev
    elif t == 3:
        formula = ev/(ev + es + nv + ns)
    elif t == 4:
        if ev < F:
            formula = 0
        elif ev == F and F != 0:
            formula = 1
        else:
            formula = 0
    elif t == 5:
        formula = ev / (ev + es + nv)
    elif t == 6:
        formula = ev / (ev + 2 * (nv + es))
    elif t == 7:
        formula = 2 * ev / (2 * ev + nv + es)
    elif t == 8:
        formula = 2 * ev / (ev + nv + es)
    elif t == 9:
        formula = (2 * ev - nv - es) / (2 * ev + nv + es)
    elif t == 10:
        formula = (ev / (ev + nv)) / (ev / (ev + nv) + es / (es + ns))
    elif t == 11:
        formula = ev / (ev + es)
    elif t == 12:
        formula = (ev / (ev + es)) - ((ev + nv) / (ev + nv + es + ns))
    elif t == 13:
        formula = ev - es
    elif t == 14:
        formula = (ev + ns - nv - es) / (ev + nv + es + ns)
    elif t == 15:
        formula = (ev + ns) / (ev + nv + es + ns)
    elif t == 16:
        formula = 2 * (ev + ns) / (2 * (ev + ns) + nv + es)
    elif t == 17:
        formula = (ev + ns) / (ev + ns + 2 * (nv + es))
    elif t == 18:
        formula = ev + ns
    elif t == 19:
        formula = np.sqrt(ev + ns)
    elif t == 20:
        formula = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / ((2 * ev + nv + es) * (2 * ns + nv + es))
    elif t == 21:
        formula = 0.5 * (ev / (2 * ev + nv + es) + ns / (2 * ns + nv + es))
    elif t == 22:
        formula = 0.5 * (ev / (ev + nv) + ev / (ev + es))
    elif t == 23:
        formula = ev / ((ev + nv) * (ev + es)) ** 0.5
    elif t == 24:
        formula = ev / (ev + ns + 2 * (nv + es))
    elif t == 25:
        formula = ev / (ev + nv) - es / (es + ns)
    elif t == 26:
        if es <= 2:
            formula = ev - es
        elif 2 < es <= 10:
            formula = ev - 2 - 0.1 * (es - 2)
        else:
            formula = ev - 2.8 - 0.001 * (es - 10)
    elif t == 27:
        formula = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + nv) + (ev + nv) * (es + ns))
    elif t == 28:
        formula = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + es) + (ev + nv) * (nv + ns))
    else:
        formula = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / (2 * ev + nv + es + 2 * ns + nv + es)

    return formula


def getSus(index, t):
    formula = 0
    try:
        formula = riskformula_new(index, t)
    except:
        if index[0] + index[1] == 0:  # e = 0
            formula = -1000
        elif index[0] + index[2] == 0:  # v = 0
            formula = -1000
        elif index[1] + index[3] == 0:  # s = 0
            formula = 1000
    return formula


def modifycase(source_case_set, MG):
    index = []
    for i in range(len(MG)):
        k = 0
        k_v = 0
        for m in range(len(MG[i])):
            if 1 in MG[i][m]:
                k += 1
            if 0 in MG[i][m]:
                k_v += 1
            if k != 0 and k_v != 0:  # 既有1又有0
                break
            else:
                k = 0
                k_v = 0
        if k == 0 or k_v == 0:  # 全0或者全1
            index.append(i)
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in index)]
    return source_case_set


def modifycase_fairness(source_case_set, MG):
    # result全为0或者全为1就删掉对应的测试用例
    index = []
    for i in range(len(MG)):
        if MG[i][0].count(0) + MG[i][0].count(3) == len(MG[i][0]) or MG[i][0].count(1) == len(MG[i][0]):
            index.append(i)
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in index)]
    return source_case_set


def modifycase_fairness2(source_case_set, MG, pf, Index):
    # result全为0或者全为1就删掉对应的测试用例
    index = []
    for i in range(len(MG)):
        if MG[i][0].count(0) + MG[i][0].count(3) == len(MG[i][0]) or MG[i][0].count(1) == len(MG[i][0]):
            index.append(i)
    source_case_set = [source_case_set[i] for i in range(len(source_case_set)) if (i not in index)]
    MG = [MG[i] for i in range(len(MG)) if (i not in index)]
    pf = [pf[i] for i in range(len(pf)) if (i not in index)]
    Index = [Index[i] for i in range(len(Index)) if (i not in index)]
    return source_case_set, MG, pf, Index


def modifycase_fairness3(MG, pf, Index):

    # result全为0或者全为1就删掉对应的测试用例
    index = []
    for i in range(len(MG)):
        if MG[i][0].count(0) + MG[i][0].count(3) == len(MG[i][0]) or MG[i][0].count(1) == len(MG[i][0]):
            index.append(i)
        # if pf[i].count(1) == len(pf[i]) or pf[i].count(0) == len(pf[i]):  #
        #     index.append(i)
    MG = [MG[i] for i in range(len(MG)) if (i not in index)]
    pf = [pf[i] for i in range(len(pf)) if (i not in index)]
    Index = [Index[i] for i in range(len(Index)) if (i not in index)]
    return MG, pf, Index


def modifycase_fairness4(MG, pf):

    # result全为0或者全为1就删掉对应的测试用例
    index = []
    for i in range(len(MG)):
        if MG[i][0].count(0) + MG[i][0].count(3) == len(MG[i][0]) or MG[i][0].count(1) == len(MG[i][0]):
            index.append(i)
        # if pf[i].count(1) == len(pf[i]) or pf[i].count(0) == len(pf[i]):  #
        #     index.append(i)
    MG = [MG[i] for i in range(len(MG)) if (i not in index)]
    pf = [pf[i] for i in range(len(pf)) if (i not in index)]
    return MG, pf


def getMetrics_rq1(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                            case = 4
                        elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                            if pf[k][i] and sus_a > sus_b:
                                Fp += 1
                                case = 1
                                # print(k, i, j, index_a, index_b, sus_a, sus_b)
                            elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                                Fp += 1
                                case = 1
                                # print(k, i, j, index_a, index_b, sus_a, sus_b)
                            else:
                                case = 3
                                Pf += 1  # p比f高

                        else:  # 只有一个failed, 但是相等
                            FP += 1
                            case = 2
                            # print(sus_a[t], sus_b[t])
                            # print(index_a, index_b)
                            # print(k, i, j)
                        cases.append(case)

            if len(cases) == 0:
                continue
            index.append(cases)

            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)
        Index.append(index)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq1_2(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        FFSlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        # case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if (pf[k][i] and sus_a > sus_b) or (pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b):
                            p1 += 1
                        if (sus_a == sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]) or \
                                (sus_a == sus_b and not pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]):
                            p2_1 += 1
                        if sus_a == sus_b and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            p2_2 += 1
                        if (sus_a > sus_b and pf[k][i * len(MG[0][0]) + j + 1] and not pf[k][i]) or \
                                (sus_a < sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]):
                            # if pf[k][i]:
                            #     p3_2 += 1
                            # else:
                            p3_1 += 1
                        # if sus_a < sus_b and pf[k][i]:
                        #     # if pf[k][i * len(MG[0][0]) + j + 1]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1


                        # if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                        #     FF += 1
                        #     # case = 4
                        # elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                        #     if pf[k][i] and sus_a > sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     else:
                        #         # case = 3
                        #         Pf += 1  # p比f高
                        #
                        # else:  # 只有一个failed, 但是相等
                        #     FP += 1
                        #     case = 2
                        # cases.append(case)

            # if len(cases) == 0:
            #     continue
            # index.append(cases)

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)

            # V_MGsum.append(V_MG)
        # Index.append(index)

        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq1_3(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        FFSlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        # case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_2 += 1
                            else:
                                p3_2 += 1
                        elif pf[k][i]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                            else:
                                p3_1 += 1
                        else:
                            if sus_a < sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                            else:
                                p3_1 += 1

                        # if (pf[k][i] and sus_a > sus_b) or (pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b):
                        #     p1 += 1
                        # if (sus_a == sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]) or \
                        #         (sus_a == sus_b and not pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]):
                        #     p2_1 += 1
                        # if sus_a == sus_b and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     p2_2 += 1
                        # if (sus_a > sus_b and pf[k][i * len(MG[0][0]) + j + 1] and not pf[k][i]) or \
                        #         (sus_a < sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]):
                        #     # if pf[k][i]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1
                        # if sus_a < sus_b and pf[k][i]:
                        #     # if pf[k][i * len(MG[0][0]) + j + 1]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1


                        # if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                        #     FF += 1
                        #     # case = 4
                        # elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                        #     if pf[k][i] and sus_a > sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     else:
                        #         # case = 3
                        #         Pf += 1  # p比f高
                        #
                        # else:  # 只有一个failed, 但是相等
                        #     FP += 1
                        #     case = 2
                        # cases.append(case)

            # if len(cases) == 0:
            #     continue
            # index.append(cases)

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)

            # V_MGsum.append(V_MG)
        # Index.append(index)

        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq1_3_new(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    统计ratio1，2，3
    Ratio1=# of VMG containing t1/# of MG containing t1 or t2
    Ratio2=# of VMG containing t2/# of MG containing t1 or t2
    Ratio3=(# of VMG containing t1+# of VMG containing t2-# of VMG containing both t1 and t2)/# of MG contain-ing t1 or t2
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'Ratio1', 'Ratio2', 'Ratio3',
                                      'VMG(%)', 'failed test case(%)', 'false satisfied MG(%)', 'fs MG / all(%)', 'fs new']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        FFSlist = []
        FFSalist = []
        FFSnewlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        r1list = []
        r2list = []
        r3list = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            r1 = []
            r2 = []
            r3 = []
            V_MG = 0
            V_MG1 = 0
            S_MG1 = 0
            ffsnew = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG
            percent_FFSa = FFS / Total_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            FFSalist.append(percent_FFSa)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                flag = 0
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        # case = 0
                        V_MG1 += 1

                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        if t == 30:
                            sus_a = nv_b
                            sus_b = nv_a
                        else:
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)

                        cta = ev_a / (sum_s + sum_v)
                        ctb = ev_b / (sum_s + sum_v)
                        ctab = (ev_a + ev_b - 1) / (sum_s + sum_v)
                        r1.append(cta)
                        r2.append(ctb)
                        r3.append(ctab)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_2 += 1
                            else:
                                p3_2 += 1
                        elif pf[k][i]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1
                            else:
                                p3_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1
                        else:
                            if sus_a < sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1
                            else:
                                p3_1 += 1
                                if flag == 0:
                                    S_MG1 += sum_s
                                    ffsnew += (MG[k][i].count(3) + MG[k][j+1].count(3))
                                else:
                                    S_MG1 += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                                    ffsnew += MG[k][j+1].count(3)
                                flag = 1

                        # if (pf[k][i] and sus_a > sus_b) or (pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b):
                        #     p1 += 1
                        # if (sus_a == sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]) or \
                        #         (sus_a == sus_b and not pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]):
                        #     p2_1 += 1
                        # if sus_a == sus_b and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     p2_2 += 1
                        # if (sus_a > sus_b and pf[k][i * len(MG[0][0]) + j + 1] and not pf[k][i]) or \
                        #         (sus_a < sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]):
                        #     # if pf[k][i]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1
                        # if sus_a < sus_b and pf[k][i]:
                        #     # if pf[k][i * len(MG[0][0]) + j + 1]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1


                        # if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                        #     FF += 1
                        #     # case = 4
                        # elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                        #     if pf[k][i] and sus_a > sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     else:
                        #         # case = 3
                        #         Pf += 1  # p比f高
                        #
                        # else:  # 只有一个failed, 但是相等
                        #     FP += 1
                        #     case = 2
                        # cases.append(case)

            # if len(cases) == 0:
            #     continue
            # index.append(cases)

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4
            if S_MG1 == 0:
                percent_ffsnew = 0
            else:
                percent_ffsnew = ffsnew / S_MG1

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)
            r1list.append(np.mean(r1))
            r2list.append(np.mean(r2))
            r3list.append(np.mean(r3))
            FFSnewlist.append(np.mean(percent_ffsnew))

            # V_MGsum.append(V_MG)
        # Index.append(index)

        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2), round(np.mean(r1list) * 100, 2),
                 round(np.mean(r2list) * 100, 2), round(np.mean(r3list) * 100, 2), round(np.mean(V_MGlist) * 100, 2), round(np.mean(Failed) * 100, 2),
                 round(np.mean(FFSlist) * 100, 2), round(np.mean(FFSalist) * 100, 2), round(np.mean(FFSnewlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq1_3_bayesian(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    统计ratio1，2，3
    Ratio1=# of VMG containing t1/# of MG containing t1 or t2
    Ratio2=# of VMG containing t2/# of MG containing t1 or t2
    Ratio3=(# of VMG containing t1+# of VMG containing t2-# of VMG containing both t1 and t2)/# of MG contain-ing t1 or t2
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss', 'Bayesian']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'Ratio1', 'Ratio2', 'Ratio3',
                                      'VMG(%)', 'failed test case(%)', 'false satisfied MG(%)', 'fs MG / all(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        FFSlist = []
        FFSalist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        r1list = []
        r2list = []
        r3list = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            r1 = []
            r2 = []
            r3 = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG
            percent_FFSa = FFS / Total_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            FFSalist.append(percent_FFSa)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        # case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        if t == 30:
                            sus_a = nv_b
                            sus_b = nv_a
                        else:
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)

                        cta = ev_a / (sum_s + sum_v)
                        ctb = ev_b / (sum_s + sum_v)
                        ctab = (ev_a + ev_b - 1) / (sum_s + sum_v)
                        r1.append(cta)
                        r2.append(ctb)
                        r3.append(ctab)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_2 += 1
                            else:
                                p3_2 += 1
                        elif pf[k][i]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                            else:
                                p3_1 += 1
                        else:
                            if sus_a < sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                            else:
                                p3_1 += 1

                        # if (pf[k][i] and sus_a > sus_b) or (pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b):
                        #     p1 += 1
                        # if (sus_a == sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]) or \
                        #         (sus_a == sus_b and not pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]):
                        #     p2_1 += 1
                        # if sus_a == sus_b and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     p2_2 += 1
                        # if (sus_a > sus_b and pf[k][i * len(MG[0][0]) + j + 1] and not pf[k][i]) or \
                        #         (sus_a < sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]):
                        #     # if pf[k][i]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1
                        # if sus_a < sus_b and pf[k][i]:
                        #     # if pf[k][i * len(MG[0][0]) + j + 1]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1


                        # if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                        #     FF += 1
                        #     # case = 4
                        # elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                        #     if pf[k][i] and sus_a > sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     else:
                        #         # case = 3
                        #         Pf += 1  # p比f高
                        #
                        # else:  # 只有一个failed, 但是相等
                        #     FP += 1
                        #     case = 2
                        # cases.append(case)

            # if len(cases) == 0:
            #     continue
            # index.append(cases)

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)
            r1list.append(np.mean(r1))
            r2list.append(np.mean(r2))
            r3list.append(np.mean(r3))

            # V_MGsum.append(V_MG)
        # Index.append(index)

        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2), round(np.mean(r1list) * 100, 2),
                 round(np.mean(r2list) * 100, 2), round(np.mean(r3list) * 100, 2), round(np.mean(V_MGlist) * 100, 2), round(np.mean(Failed) * 100, 2),
                 round(np.mean(FFSlist) * 100, 2), round(np.mean(FFSalist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq1_4(row, ws, mu, MG, pf):
    '''
    liuhuai
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        FFSlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_1_1 = 0
            p2_1_2 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        # case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_2 += 1
                            else:
                                p3_2 += 1
                        elif pf[k][i]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1_1 += 1
                            else:
                                p3_1 += 1
                        else:
                            if sus_a < sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1_2 += 1
                            else:
                                p3_1 += 1




                        # if (pf[k][i] and sus_a > sus_b) or (pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b):
                        #     p1 += 1
                        # if (sus_a == sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]) or \
                        #         (sus_a == sus_b and not pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]):
                        #     p2_1 += 1
                        # if sus_a == sus_b and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     p2_2 += 1
                        # if (sus_a > sus_b and pf[k][i * len(MG[0][0]) + j + 1] and not pf[k][i]) or \
                        #         (sus_a < sus_b and pf[k][i] and not pf[k][i * len(MG[0][0]) + j + 1]):
                        #     # if pf[k][i]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1
                        # if sus_a < sus_b and pf[k][i]:
                        #     # if pf[k][i * len(MG[0][0]) + j + 1]:
                        #     #     p3_2 += 1
                        #     # else:
                        #     p3_1 += 1


                        # if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        #     # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                        #     FF += 1
                        #     # case = 4
                        # elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                        #     if pf[k][i] and sus_a > sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                        #         Fp += 1
                        #         # case = 1
                        #         # print(k, i, j, index_a, index_b, sus_a, sus_b)
                        #     else:
                        #         # case = 3
                        #         Pf += 1  # p比f高
                        #
                        # else:  # 只有一个failed, 但是相等
                        #     FP += 1
                        #     case = 2
                        # cases.append(case)
            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        sum_s = MG[k][i].count(0) + MG[k][i].count(3)
                        sum_v = MG[k][i].count(1)
                        sum_s += (MG[k][j+1].count(0) + MG[k][j+1].count(3))
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0) + MG[k][i].count(3)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0) + MG[k][j+1].count(3)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)
                        if sus_a == sus_b and not (pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]):
                            if (p2_1_1 > p2_1_2 and pf[k][i]) or (p2_1_1 < p2_1_2 and pf[k][i * len(MG[0][0]) + j + 1]):
                                p1 += 1
                            elif p2_1_1 == p2_1_2:
                                p2_1 += 1
                            else:
                                p3_1 += 1


            # if len(cases) == 0:
            #     continue
            # index.append(cases)

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)

            # V_MGsum.append(V_MG)
        # Index.append(index)

        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq2(row, ws, mu, MG, pf, d):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):  # len(Formula)
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        Pfcount = 0
        V_MG1count = 0
        for k in range(len(MG)):  # len(MG)
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            index = []
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                        # print(d, k, i, j)
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            index.append(i)
                            index.append(i * len(MG[0][0]) + j + 1)
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue
            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # 因为需要两个测试用例同时超过k次, 所以只能考虑第一行的MG
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0)
                        sum_v = MG[k][i].count(1)
                        sum_s += MG[k][j+1].count(0)
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)
                        if ev_a + es_a > d and ev_b + es_b > d:
                            V_MG1 += 1

                            if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                                # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                                FF += 1
                            elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                                if pf[k][i] and sus_a > sus_b:
                                    Fp += 1
                                elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                                    Fp += 1
                                else:
                                    Pf += 1  # p比f高
                                    index.append((k, i, j))

                            else:  # 只有一个failed, 但是相等
                                FP += 1
                                # print(sus_a[t], sus_b[t])
                                # print(index_a, index_b)
                                # print(k, i, j)
            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)


            # V_MGsum.append(V_MG)
        # print(Pflist)
        # print(round(np.mean(Pflist) * 100, 2))
        # print(t, Pfcount, V_MG1count)
        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    pd = np.array(datadist['Naish1'])
    if np.isnan(pd[0]):
        print("Mutant" + str(mu) + '没有数值')
    else:
        for i, j in datadist.items():  # i--公式名称, j--指标值
            ws.cell(row, 1).value = i  # 添加第 1 列的数据
            for col in range(2, len(j) + 2):  # values列表中索引
                ws.cell(row, col).value = j[col - 2]
            row += 1  # 行数
        row += 2  # 行数
    return row


def getMetrics_rq2_3(row, ws, mu, MG, pf, d):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'VMG(%)', 'VMG(#)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):  # len(Formula)
        V_MGlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        FFSlist = []
        Failed = []
        V_MG1sum = []
        Pfcount = 0
        V_MG1count = 0
        for k in range(len(MG)):  # len(MG)
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            index = []
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                        # print(d, k, i, j)
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            index.append(i)
                            index.append(i * len(MG[0][0]) + j + 1)
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue
            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # 因为需要两个测试用例同时超过k次, 所以只能考虑第一行的MG
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0)
                        sum_v = MG[k][i].count(1)
                        sum_s += MG[k][j+1].count(0)
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)
                        if (ev_a + es_a) > d and (ev_b + es_b) > d:
                            if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                                print('出错了：p+p=v')
                            V_MG1 += 1
                            if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                                if sus_a > sus_b:
                                    p1 += 1
                                elif sus_a == sus_b:
                                    p2_2 += 1
                                else:
                                    p3_2 += 1
                            elif pf[k][i]:
                                if sus_a > sus_b:
                                    p1 += 1
                                elif sus_a == sus_b:
                                    p2_1 += 1
                                else:
                                    p3_1 += 1
                                    # print(d,k,i,j)
                            else:
                                if sus_a < sus_b:
                                    p1 += 1
                                elif sus_a == sus_b:
                                    p2_1 += 1
                                else:
                                    p3_1 += 1
                                    # print(d, k, i, j)

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)


            V_MG1sum.append(V_MG1)
        # print(Pflist)
        # print(round(np.mean(Pflist) * 100, 2))
        # print(t, Pfcount, V_MG1count)
        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2), round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2), round(np.mean(V_MGlist) * 100, 2), np.sum(V_MG1sum),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    pd = np.array(datadist['Naish1'])
    if np.isnan(pd[0]):
        print("Mutant" + str(mu) + '没有数值')
    else:
        for i, j in datadist.items():  # i--公式名称, j--指标值
            ws.cell(row, 1).value = i  # 添加第 1 列的数据
            for col in range(2, len(j) + 2):  # values列表中索引
                ws.cell(row, col).value = j[col - 2]
            row += 1  # 行数
        row += 2  # 行数
    return row


def getMetrics_rq3(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    去除fs，带fs的见rq1
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0)
                        sum_v = MG[k][i].count(1)
                        sum_s += MG[k][j+1].count(0)
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                            case = 4
                        elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                            if pf[k][i] and sus_a > sus_b:
                                Fp += 1
                                case = 1
                                # print(k, i, j, index_a, index_b, sus_a, sus_b)
                            elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                                Fp += 1
                                case = 1
                                # print(k, i, j, index_a, index_b, sus_a, sus_b)
                            else:
                                case = 3
                                Pf += 1  # p比f高

                        else:  # 只有一个failed, 但是相等
                            FP += 1
                            case = 2
                            # print(sus_a[t], sus_b[t])
                            # print(index_a, index_b)
                            # print(k, i, j)
                        cases.append(case)

            if len(cases) == 0:
                continue
            index.append(cases)

            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)
        Index.append(index)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_rq3_3(row, ws, mu, MG, pf):
    '''
    根据TY改, 统计第一行
    去除fs，带fs的见rq1
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['p1(%)', 'p2_1(%)', 'p2_2(%)', 'p3_1(%)', 'p3_2(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        p1list = []
        p2_1list = []
        p2_2list = []
        p3_1list = []
        p3_2list = []
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):  # k代表source的个数
            cases = []
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            p1 = 0
            p2_1 = 0
            p2_2 = 0
            p3_1 = 0
            p3_2 = 0
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 3:
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        if not (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            print('出错了：p+p=v')
                        case = 0
                        V_MG1 += 1
                        # 根据包含的测试用例确定 MG set
                        sum_s = MG[k][i].count(0)
                        sum_v = MG[k][i].count(1)
                        sum_s += MG[k][j+1].count(0)
                        sum_v += MG[k][j+1].count(1)
                        # 求每个测试用例的ev es nv ns
                        ev_a = MG[k][i].count(1)
                        es_a = MG[k][i].count(0)
                        nv_a = sum_v - ev_a
                        ns_a = sum_s - es_a
                        index_a = [ev_a, es_a, nv_a, ns_a]

                        ev_b = MG[k][j+1].count(1) + 1
                        es_b = MG[k][j+1].count(0)
                        nv_b = sum_v - ev_b
                        ns_b = sum_s - es_b
                        index_b = [ev_b, es_b, nv_b, ns_b]
                        # 求测试用例的可疑度
                        sus_a = getSus(index_a, t)
                        sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_2 += 1
                            else:
                                p3_2 += 1
                        elif pf[k][i]:
                            if sus_a > sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                            else:
                                p3_1 += 1
                        else:
                            if sus_a < sus_b:
                                p1 += 1
                            elif sus_a == sus_b:
                                p2_1 += 1
                            else:
                                p3_1 += 1

            if V_MG1 == 0:
                continue
            percent_p1 = p1 / V_MG1  # 1
            percent_p2_1 = p2_1 / V_MG1  # 2
            percent_p2_2 = p2_2 / V_MG1  # 3
            percent_p3_1 = p3_1 / V_MG1  # 4
            percent_p3_2 = p3_2 / V_MG1  # 4

            p1list.append(percent_p1)
            p2_1list.append(percent_p2_1)
            p2_2list.append(percent_p2_2)
            p3_1list.append(percent_p3_1)
            p3_2list.append(percent_p3_2)

            # V_MGsum.append(V_MG)
        # Index.append(index)

        value = [round(np.mean(p1list) * 100, 2), round(np.mean(p2_1list) * 100, 2),
                 round(np.mean(p2_2list) * 100, 2),
                 round(np.mean(p3_1list) * 100, 2), round(np.mean(p3_2list) * 100, 2),
                 round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics(row, ws, mu, MG, Risk, pf):
    '''
    统计所有MG
    '''
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):
            V_MG = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(len(MG[k])):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        # if k == 1 and i == 1 and j == 0:
                        #     print(t, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1], Risk[k][i][t], Risk[k][i * len(MG[0][0]) + j + 1][t])
                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                        elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                            if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
                                    Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            else:
                                Pf += 1  # p比f高
                                index.append((k, i, j))
                                # print(i)
                                # print(k, i)
                                # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        else:  # 只有一个failed, 但是相等
                            FP += 1

            percent_Fp = Fp / V_MG  # 1
            percent_FP = FP / V_MG  # 2
            percent_Pf = Pf / V_MG  # 3
            percent_FF = FF / V_MG  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
        Index.append(index)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_v2(row, ws, mu, MG, Risk, pf):
    '''
    根据TY改, 统计第一行
    '''
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):
            cases = []
            V_MG = 0
            S_MG = 0
            V_MG1 = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG1 += 1
                        case = 0
                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                            case = 4
                        elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                            if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                                case = 1
                                # print(k, i, j, Risk[k][i][t], Risk[k][i * len(MG[0][0]) + j + 1][t])
                            elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
                                    Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                                case = 1
                                # print(k, i, j, Risk[k][i][t], Risk[k][i * len(MG[0][0]) + j + 1][t])
                            else:
                                Pf += 1  # p比f高
                                case = 3
                                # print(t, k, i, j)
                                # print(k, i)
                                # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        else:  # 只有一个failed, 但是相等
                            FP += 1
                            case = 2
                        cases.append(case)
            if len(cases) == 0:
                continue
            index.append(cases)

            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)
        Index.append(index)
        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row, Index


def getMetrics_new(row, ws, mu, MG, pf):
    '''
    根据TY改
    '''

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)): # k代表source的个数
            V_MG = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            FFS += 1

            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(len(MG[k])):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        # 根据包含的测试用例确定 MG set
                        if i == 0:
                            sum_s = MG[k][i].count(0)
                            sum_v = MG[k][i].count(1)
                            sum_s += MG[k][j+1].count(0)
                            sum_v += MG[k][j+1].count(1)
                            # 求每个测试用例的ev es nv ns
                            ev_a = MG[k][i].count(1)
                            es_a = MG[k][i].count(0)
                            nv_a = sum_v - ev_a
                            ns_a = sum_s - es_a
                            index_a = [ev_a, es_a, nv_a, ns_a]

                            ev_b = MG[k][j+1].count(1) + 1
                            es_b = MG[k][j+1].count(0)
                            nv_b = sum_v - ev_b
                            ns_b = sum_s - es_b
                            index_b = [ev_b, es_b, nv_b, ns_b]
                            # 求测试用例的可疑度
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)
                        else:
                            sum_s = MG[k][i].count(0)
                            sum_v = MG[k][i].count(1)
                            ev_a = MG[k][i].count(1)
                            es_a = MG[k][i].count(0)
                            if MG[k][0][i-1] == 0:
                                sum_s += 1
                                ns_a = sum_s - es_a
                                nv_a = sum_v - ev_a
                                es_a += 1
                                ns_a -= 1
                            else:
                                sum_v += 1
                                ns_a = sum_s - es_a
                                nv_a = sum_v - ev_a
                                ev_a += 1
                                nv_a -= 1
                            ev_b = 1
                            nv_b = sum_v - ev_b
                            es_b = 0
                            ns_b = sum_s
                            index_a = [ev_a, es_a, nv_a, ns_a]
                            index_b = [ev_b, es_b, nv_b, ns_b]
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                        elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                            if pf[k][i] and sus_a > sus_b:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                                Fp += 1
                            else:
                                Pf += 1  # p比f高
                                index.append((k, i, j))
                                # print(i)

                        else:  # 只有一个failed, 但是相等
                            FP += 1
                            # print(sus_a[t], sus_b[t])
                            # print(index_a, index_b)
                            # print(k, i, j)

            percent_Fp = Fp / V_MG  # 1
            percent_FP = FP / V_MG  # 2
            percent_Pf = Pf / V_MG  # 3
            percent_FF = FF / V_MG  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
        Index.append(index)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


# def getMetrics_v2(row, ws, mu, MG, Risk, pf):
#
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
#                'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
#                'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
#                                       'failed test case(%)', 'false satisfied MG(%)']}
#     datadist.update(tablelist)
#     Index = []
#     for t in range(len(Formula)):
#         V_MGlist = []
#         Fplist = []  # Case 1: t1(f) > t2(p)
#         FPlist = []  # Case 2: t1(f) = t2(p)
#         Pflist = []  # Case 3: t1(f) < t2(p)
#         FFlist = []  # Case 4: t1 and t2 both fail
#         FFSlist = []
#         Failed = []
#         V_MGsum = []
#         index = []
#         for k in range(len(MG)):
#             V_MG = 0
#             S_MG = 0
#             Fp = 0  # Case 1: t1(f) > t2(p)
#             FP = 0  # Case 2: t1(f) = t2(p)
#             Pf = 0  # Case 3: t1(f) < t2(p)
#             FF = 0  # Case 4: t1 and t2 both fail
#             FFS = 0
#             Total_MG = len(MG[0]) * len(MG[0][0])
#             failedcase = pf[k].count(1)
#             percent_failed = failedcase / len(pf[k])
#             for i in range(len(MG[k])):
#                 for j in range(len(MG[k][i])):
#                     if MG[k][i][j] == 1:  # violated
#                         V_MG += 1
#                     elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
#                         S_MG += 1
#                         if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
#                             FFS += 1
#             percent_VMG = V_MG / Total_MG
#             percent_FFS = FFS / S_MG
#
#             if V_MG == 0:
#                 continue
#
#             V_MGlist.append(percent_VMG)
#             FFSlist.append(percent_FFS)
#             Failed.append(percent_failed)
#
#             for i in range(len(MG[k])):  # len(MG[k])
#                 for j in range(len(MG[k][i])):
#                     if MG[k][i][j] == 1:  # violated
#                         # if k == 1 and i == 1 and j == 0:
#                         #     print(t, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1], Risk[k][i][t], Risk[k][i * len(MG[0][0]) + j + 1][t])
#                         if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
#                             # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
#                             FF += 1
#                         elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
#                             if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
#                                 Fp += 1
#                             elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
#                                     Risk[k][i * len(MG[0][0]) + j + 1][t]:
#                                 Fp += 1
#                             else:
#                                 Pf += 1  # p比f高
#                                 index.append((k, i, j))
#                                 # print(t, k, i, j)
#                                 # print(k, i)
#                                 # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
#                         else:  # 只有一个failed, 但是相等
#                             FP += 1
#
#             percent_Fp = Fp / V_MG  # 1
#             percent_FP = FP / V_MG  # 2
#             percent_Pf = Pf / V_MG  # 3
#             percent_FF = FF / V_MG  # 4
#             Fplist.append(percent_Fp)
#             FPlist.append(percent_FP)
#             Pflist.append(percent_Pf)
#             FFlist.append(percent_FF)
#
#             # V_MGsum.append(V_MG)
#
#         value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
#                  round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
#                  round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         Index.append(index)
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row


def getMetrics_mored_v2(row, ws, mu, MG, Risk, pf, d, Index):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        Pfcount = 0
        V_MG1count = 0
        for k in range(len(MG)):
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            index = []
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            index.append(i)
                            index.append(i * len(MG[0][0]) + j + 1)
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1 and \
                            Index[k][i * len(MG[0][0]) + j + 1][0] + Index[k][i * len(MG[0][0]) + j + 1][1] > d and \
                            Index[k][i][0] + Index[k][i][1] > d:  # violated
                        V_MG1 += 1
                        # print(k, i, j)
                        V_MG1count += 1
                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                        elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                            if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
                                    Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            else:
                                Pf += 1  # p比f高
                                Pfcount += 1
                                # if i in index or i * len(MG[0][0]) + j + 1 in index:
                                #     print(True)
                                # print(k, i)
                                # print(t, k, i, j)
                                # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        else:  # 只有一个failed, 但是相等
                            FP += 1
            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)
        # print(Pflist)
        # print(round(np.mean(Pflist) * 100, 2))
        # print(t, Pfcount, V_MG1count)
        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    pd = np.array(datadist['Naish1'])
    if np.isnan(pd[0]):
        print("Mutant" + str(mu) + '没有数值')
    else:
        for i, j in datadist.items():  # i--公式名称, j--指标值
            ws.cell(row, 1).value = i  # 添加第 1 列的数据
            for col in range(2, len(j) + 2):  # values列表中索引
                ws.cell(row, col).value = j[col - 2]
            row += 1  # 行数
        row += 2  # 行数
    return row


def getMetrics_v2_new(row, ws, mu, MG, pf):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    Index = []
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        index = []
        for k in range(len(MG)):
            V_MG = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG
            if V_MG == 0:
                continue

            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)

            for i in range(len(MG[k])):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        # 根据包含的测试用例确定 MG set
                        if i == 0:
                            sum_s = MG[k][i].count(0)
                            sum_v = MG[k][i].count(1)
                            sum_s += MG[k][j+1].count(0)
                            sum_v += MG[k][j+1].count(1)
                            # 求每个测试用例的ev es nv ns
                            ev_a = MG[k][i].count(1)
                            es_a = MG[k][i].count(0)
                            nv_a = sum_v - ev_a
                            ns_a = sum_s - es_a
                            index_a = [ev_a, es_a, nv_a, ns_a]

                            ev_b = MG[k][j+1].count(1) + 1
                            es_b = MG[k][j+1].count(0)
                            nv_b = sum_v - ev_b
                            ns_b = sum_s - es_b
                            index_b = [ev_b, es_b, nv_b, ns_b]
                            # 求测试用例的可疑度
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)
                        else:
                            sum_s = MG[k][i].count(0)
                            sum_v = MG[k][i].count(1)
                            ev_a = MG[k][i].count(1)
                            es_a = MG[k][i].count(0)
                            if MG[k][0][i-1] == 0:
                                sum_s += 1
                                ns_a = sum_s - es_a
                                nv_a = sum_v - ev_a
                                es_a += 1
                                ns_a -= 1
                            else:
                                sum_v += 1
                                ns_a = sum_s - es_a
                                nv_a = sum_v - ev_a
                                ev_a += 1
                                nv_a -= 1
                            ev_b = 1
                            nv_b = sum_v - ev_b
                            es_b = 0
                            ns_b = sum_s
                            index_a = [ev_a, es_a, nv_a, ns_a]
                            index_b = [ev_b, es_b, nv_b, ns_b]
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)

                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                        elif not (sus_a == sus_b):  # 只有一个failed, 并且可疑度不相等
                            if pf[k][i] and sus_a > sus_b:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and sus_a < sus_b:
                                Fp += 1
                            else:
                                Pf += 1  # p比f高
                                index.append((k, i, j))

                        else:  # 只有一个failed, 但是相等
                            FP += 1
                            # print(sus_a[t], sus_b[t])
                            # print(index_a, index_b)
                            # print(k, i, j)

            percent_Fp = Fp / V_MG  # 1
            percent_FP = FP / V_MG  # 2
            percent_Pf = Pf / V_MG  # 3
            percent_FF = FF / V_MG  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)

            # V_MGsum.append(V_MG)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
        Index.append(index)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_fairness(row, ws, mu, MG, risk, pf, pf_new):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)', 'failed test case(%)',
                                      'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        for i in range(len(MG)):
            V_MG = 0
            V_MG_s = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            testcase = len(pf_new[i]) * len(pf_new[i][1]) + 1
            failedcase = 0
            Total_MG = len(MG[i]) * len(MG[i][1]) + 1
            for k in range(len(pf_new[i])):
                failedcase += pf_new[i][k].count(1)
            percent_failed = failedcase / testcase
            for j in range(len(MG[i][0])):
                if MG[i][0][j]:  # violated MG
                    V_MG_s += 1
                    # 统计case
                    if pf[i][0] and pf[i][j + 1]:
                        # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                        FF += 1
                    elif not (risk[i * len(pf[0])][t] == risk[i * len(pf[0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                        if pf[i][0] and risk[i * len(pf[0])][t] > risk[i * len(pf[0]) + j + 1][t]:
                            Fp += 1
                        elif pf[i][j + 1] and risk[i * len(pf[0])][t] < risk[i * len(pf[0]) + j + 1][t]:
                            Fp += 1
                        else:
                            Pf += 1  # p比f高
                            # print((pf[i][0],risk[i*len(pf[0])][t]),(pf[i][j + 1],risk[i * len(pf[0]) + j + 1][t]))
                    else:  # 只有一个failed, 但是相等
                        FP += 1
            for m in range(len(MG[i])):
                V_MG += MG[i][m].count(1)
                for n in range(len(MG[i][m])):
                    if not MG[i][m][n]:
                        S_MG += 1
                        if pf_new[i][m][0] and pf_new[i][m][n + 1]:
                            FFS += 1

            percent_Fp = Fp / V_MG_s  # 1
            percent_FP = FP / V_MG_s  # 2
            percent_Pf = Pf / V_MG_s  # 3
            percent_FF = FF / V_MG_s  # 4
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)
            Failed.append(percent_failed)
            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)

        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数

    return row


def getMetrics_moreone(row, ws, mu, MG, Risk, pf):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):  # len(Formula)
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        for k in range(len(MG)):
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0:
                        S_MG += 1
                        if pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]:
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG1 += 1
                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                        elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                            if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
                                    Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            else:
                                Pf += 1  # p比f高
                                # if i in index or i * len(MG[0][0]) + j + 1 in index:
                                #     print(True)
                                # print(k, i)
                                # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        else:  # 只有一个failed, 但是相等
                            FP += 1
            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)
            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)
            # V_MGsum.append(V_MG)
        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_nofalse(row, ws, mu, MG, Risk, pf):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        FSlist = []
        V_MGsum = []
        s_VMG = 0
        for k in range(len(MG)):
            index = []
            V_MG = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            FS = 0 # Case 1 and Case 4中包含fs的 MG的比例
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
                        S_MG += 1
                        if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
                            FFS += 1
                            # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        if MG[k][i][j] == 3:
                            index.append(i)
                            index.append(i * len(MG[0][0]) + j + 1)
            percent_FFS = FFS / S_MG

            for i in range(len(MG[k])):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    # if MG[k][i][j] and (i in index or (i * len(MG[0][0]) + j + 1) in index):
                    #     V_MG -= 1
                    #     continue
                    if MG[k][i][j] == 1:  # violated
                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                            if i in index or (i * len(MG[0][0]) + j + 1) in index:
                                FS += 1
                        elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                            if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
                                    Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            else:
                                # if i in index or (i * len(MG[0][0]) + j + 1) in index:
                                #     V_MG -= 1
                                #     continue
                                # else:
                                Pf += 1  # p比f高
                                if i in index or (i * len(MG[0][0]) + j + 1) in index:
                                    FS += 1
                                # print(k, i, j)
                                # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        else:  # 只有一个failed, 但是相等
                            FP += 1

            percent_Fp = Fp / V_MG  # 1
            percent_FP = FP / V_MG  # 2
            percent_Pf = Pf / V_MG  # 3
            percent_FF = FF / V_MG  # 4
            percent_VMG = V_MG / Total_MG
            percent_FS = FS / V_MG
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)
            FFlist.append(percent_FF)
            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)
            FSlist.append(percent_FS)
        # if len(Fplist) == 0:
            # V_MGsum.append(V_MG)
            # s_VMG += V_MG
        # print(s_VMG)
        # print(round(np.mean(FSlist) * 100, 2))
        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


def getMetrics_mored(row, ws, mu, MG, Risk, pf, d, Index):

    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice', 'Goodman',
               'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto', 'Hamming etc.',
               'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
                                      'failed test case(%)', 'false satisfied MG(%)']}
    datadist.update(tablelist)
    for t in range(len(Formula)):
        V_MGlist = []
        Fplist = []  # Case 1: t1(f) > t2(p)
        FPlist = []  # Case 2: t1(f) = t2(p)
        Pflist = []  # Case 3: t1(f) < t2(p)
        FFlist = []  # Case 4: t1 and t2 both fail
        FFSlist = []
        Failed = []
        V_MGsum = []
        Pfcount = 0
        V_MG1count = 0
        for k in range(len(MG)):
            V_MG = 0
            V_MG1 = 0
            S_MG = 0
            Fp = 0  # Case 1: t1(f) > t2(p)
            FP = 0  # Case 2: t1(f) = t2(p)
            Pf = 0  # Case 3: t1(f) < t2(p)
            FF = 0  # Case 4: t1 and t2 both fail
            FFS = 0
            Total_MG = len(MG[0]) * len(MG[0][0])
            failedcase = pf[k].count(1)
            percent_failed = failedcase / len(pf[k])
            index = []
            for i in range(len(MG[k])):
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1:  # violated
                        V_MG += 1
                    elif MG[k][i][j] == 0:
                        S_MG += 1
                        if pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]:
                            index.append(i)
                            index.append(i * len(MG[0][0]) + j + 1)
                            FFS += 1
            percent_VMG = V_MG / Total_MG
            percent_FFS = FFS / S_MG

            for i in range(1):  # len(MG[k])
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1 and \
                            Index[k][i * len(MG[0][0]) + j + 1][0] + Index[k][i * len(MG[0][0]) + j + 1][1] >= d:  # violated
                        V_MG1 += 1
                        # print(k, i, j)
                        V_MG1count += 1
                        if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                            # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
                            FF += 1
                        elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
                            if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
                                    Risk[k][i * len(MG[0][0]) + j + 1][t]:
                                Fp += 1
                            else:
                                Pf += 1  # p比f高
                                Pfcount += 1
                                # if i in index or i * len(MG[0][0]) + j + 1 in index:
                                #     print(True)
                                # print(k, i)
                                # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
                        else:  # 只有一个failed, 但是相等
                            FP += 1
            if V_MG1 == 0:
                continue
            percent_Fp = Fp / V_MG1  # 1
            percent_FP = FP / V_MG1  # 2
            percent_Pf = Pf / V_MG1  # 3
            percent_FF = FF / V_MG1  # 4
            Fplist.append(percent_Fp)
            FPlist.append(percent_FP)
            Pflist.append(percent_Pf)

            FFlist.append(percent_FF)
            V_MGlist.append(percent_VMG)
            FFSlist.append(percent_FFS)
            Failed.append(percent_failed)
            # V_MGsum.append(V_MG)
        # print(Pflist)
        # print(round(np.mean(Pflist) * 100, 2))
        # print(Pfcount, V_MG1count)
        value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
                 round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
                 round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row