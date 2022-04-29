'''
RQ1
Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
from KNN import *
import math
import numpy as np
import copy
import random
import json
import csv
from publicFun import *
from openpyxl import load_workbook

random.seed(1)


def getSource(dynamic):
    Follow = []
    source_case_set = []
    for k in range(200):
        trainingSet = random.sample(dataset, int(len(dataset) * 0.6))
        testSet = copy.deepcopy(dataset)
        for i in range(len(trainingSet)):
            testSet.remove(trainingSet[i])  # 被测样本
        testSet = random.sample(testSet, 1)
        source_case_set.append([trainingSet, testSet])

    for i in range(len(source_case_set)):
        MGS, follow_case = MTG(source_case_set[i], dynamic)
        Follow.append(follow_case)

    MG = []
    for i in range(len(source_case_set)):
        MGS, Result = riskIndex(source_case_set[i], dynamic)
        MG.append(MGS)
    source_case_set = modifycase_fairness(source_case_set, MG)

    if len(source_case_set) == 0:
        print('测试用例集不满足要求')

    return source_case_set


def riskIndex(argv, dynamic):
    MGS = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        knn = KNN()
        knn.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_a = knn.getPredications()
        dynamic.setInput(testcase[i][0], testcase[i][1])  # oracle
        result_s_m = dynamic.getPredications()
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    # p or f统计完
    # 统计全部的v和s

    return MGS, Result


# def getRisk(source_case_set):
#     pf = []
#     MG = []
#     Risk = []
#     for i in range(len(source_case_set)):
#         risk = []
#         MGS, index, Result = riskIndex(source_case_set[i], dynamic)
#         pf.append(Result)
#         MG.append(MGS)
#         for j in range(len(index)):
#             formula = riskformula(index[j])
#             risk.append(formula)
#         Risk.append(risk)
#     return pf, MG, Risk

if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/correlation_with3.xlsx'
    wb = load_workbook(path)
    string = 'knn'
    del wb[string]
    ws = wb.create_sheet(string)
    with open('/Applications/work/data/MT/FAILTIM/Result/knn/iris.data', 'rt') as csvfile:
        lines = csv.reader(csvfile)
        dataset = list(lines)
    MG_set = []
    source = []
    for mu in range(1, 11):
        # dynamic = KNNFactory("MU_" + str(mu) + "_KNN").getKNN()
        # source_case_set = getSource(dynamic)
        # # data = {
        # #     'source_case_set': source_case_set
        # # }
        # # # 将数据存下来
        # # json_str = json.dumps(data)
        # # with open('Result/KNN/mutant' + str(mu) + '_rq1.json', 'w') as f:
        # #     json.dump(json_str, f)
        # # pf, MG, Risk = getRisk(source_case_set)
        # with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # source_case_set = data['source_case_set']
        # source.append(source_case_set)
        # MG_set = []
        # pf_set = []
        # for i in range(len(source_case_set)):
        #     MG, pf = riskIndex(source_case_set[i], dynamic)
        #     MG_set.append(MG)
        #     pf_set.append(pf)
        # data = {
        #         'source_case_set': source_case_set, 'pf': pf_set, 'MG': MG_set
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('Result/KNN/mutant' + str(mu) + '_rq1.json', 'w') as f:
        #     json.dump(json_str, f)
        # pf_set = data['pf']
        # MG_set = data['MG']

        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']
        MG_set.append(MG)
        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue
        row = getMetrics_rq1_3_new(row, ws, mu, MG, pf)

        # data = {
        #     'cases': cases
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open(
        #         '/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_newcases.json',
        #         'w') as f:
        #     json.dump(json_str, f)

    # wb.save(path)

