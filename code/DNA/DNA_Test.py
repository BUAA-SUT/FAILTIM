import random
import re
from Execution import *
from TestCase import *
import itertools
from openpyxl import load_workbook
import copy
import sys
import time
import json
from DNA import *
import sys
sys.path.append('../')
from PublicFun import *


def getSource(mr_list, test_case, num_of_samples):
    # for i in range(num_of_samples):
    #     test_case.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
    #     test_case.generateRandomTestcase()
    for i in range(num_of_samples):
        for j in range(5, len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
            mr.getFollow(j)
    for i in range(num_of_samples):
        for j in range(len(mr_list)):
            if j < 5:
                for k in range(5, len(mr_list)):
                    mr = mr_list[k]
                    mr.setTestCase(test_case)
                    mr.original_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}".format(i), "outtree_{}".format(i))
                    mr.getFollow(k)
            else:
                for k in range(len(mr_list)):
                    mr = mr_list[k]
                    mr.setTestCase(test_case)
                    mr.original_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}".format(i), "outtree_{}".format(i))
                    mr.getFollow(k)


def getMG(mu, ts, num_of_samples, mr_list):
    MGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("infile_{}".format(i), "outfile_{}_{}".format(mu, i), "outtree_{}_{}".format(mu, i))
        # ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
        original_output = MR().getResults(ts)
        followup_ts = ts
        for j in range(len(mr_list)):
            # followup_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}".format(i, j), "outtree_{}_{}".format(i, j))
            followup_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}_{}".format(mu, i, j),
                                        "outtree_{}_{}_{}".format(mu, i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts)
            expected_output = mr.getExpectedOutput(original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("infile_{}_{}_f".format(i, m), "outfile_{}_{}_{}".format(mu, i, m),
                                           "outtree_{}_{}_{}".format(mu, i, m))
            # ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
            original_output = MR().getResults(ts)
            followup_ts = ts
            for n in range(len(mr_list)):
                # followup_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}".format(i, j), "outtree_{}_{}".format(i, j))
                followup_ts.setInputOutput("infile_{}_{}_f_{}_f".format(i, m, n), "outfile_{}_{}_{}_{}".format(mu, i, m, n),
                                           "outtree_{}_{}_{}_{}".format(mu, i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts)
                expected_output = mr.getExpectedOutput(original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
            MGs.append(MG)
        MGS.append(MGs)
    return MGS

def getpf(mu, ts, num_of_samples, mr_list):
    pf = []
    for i in range(num_of_samples):
        result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
        ts.setInputOutput("infile_{}".format(i), "outfile_{}_{}".format(0, i), "outtree_{}_{}".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("infile_{}".format(i), "outfile_{}_{}".format(mu, i), "outtree_{}_{}".format(mu, i))
        program_output = MR().getResults(program_ts)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result[0] = 1
        for j in range(len(mr_list)):
            ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}_{}".format(0, i, j),
                                        "outtree_{}_{}_{}".format(0, i, j))
            original_output = MR().getResults(ts)
            program_ts = ts
            program_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}_{}".format(mu, i, j),
                                        "outtree_{}_{}_{}".format(mu, i, j))
            program_output = MR().getResults(program_ts)
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result[j + 1] = 1

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("infile_{}_{}_f_{}_f".format(i, m, n), "outfile_{}_{}_{}_{}".format(0, i, m, n),
                                           "outtree_{}_{}_{}_{}".format(0, i, m, n))
                original_output = MR().getResults(ts)
                program_ts = ts
                program_ts.setInputOutput("infile_{}_{}_f_{}_f".format(i, m, n), "outfile_{}_{}_{}_{}".format(mu, i, m, n),
                                           "outtree_{}_{}_{}_{}".format(mu, i, m, n))
                program_output = MR().getResults(program_ts)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
        pf.append(result)
    return pf

def riskIndex(MGS):
    index = []
    # 统计全部的v和s
    sum_s = 0
    sum_v = 0
    for i in range(len(MGS)):
        sum_s += MGS[i].count(0)
        sum_v += MGS[i].count(1)
    if not sum_s + sum_v == (len(MGS) * len(MGS[0])):
        print('异常')

    for i in range((len(MGS[0]) * (len(MGS[0]) + 1) + 1)):
        if i == 0:
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
        elif i <= len(MGS[0]):
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
            if MGS[0][i - 1] == 0:
                es += 1
                ns -= 1
            else:
                ev += 1
                nv -= 1
        else:
            t = int((i - 1) / len(MGS[0]))
            if MGS[t][(i - 1) % len(MGS[0])] == 0:
                es = 1
                ns = sum_s - es
                ev = 0
                nv = sum_v
            else:
                ev = 1
                nv = sum_v - ev
                es = 0
                ns = sum_s
        index.append([ev, es, nv, ns])
    return index



def getIndex(MG, pf):
    MG, pf = modifycase_fairness2(MG, pf)
    return MG, pf


if __name__ == "__main__":
    myenv = MyEnv()
    myenv.CreateWorkingDirs()
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/result1.xlsx'
    wb = load_workbook(path)
    string = 'DNA'
    del wb[string]
    ws = wb.create_sheet(string)
    ts = TestCase()
    num_of_samples = 100  # 测试用例个数
    cmr1 = CompositionMR()
    cmr2 = CompositionMR()
    cmr3 = CompositionMR()
    cmr4 = CompositionMR()
    cmr5 = CompositionMR()
    cmr6 = CompositionMR()
    cmr1.setMRs([MR1(), MR2()])
    cmr2.setMRs([MR2(), MR1()])
    cmr3.setMRs([MR1(), MR3()])
    cmr4.setMRs([MR3(), MR1()])
    cmr5.setMRs([MR2(), MR3()])
    cmr6.setMRs([MR3(), MR2()])
    mr_list = [MR1(), MR2(), MR3(), MR4(), MR6(), cmr1, cmr2, cmr3, cmr4, cmr5, cmr6]
    mutants_list = ["v0.exe", "v1.exe", "v2.exe", "v3.exe", "v4.exe", "v5.exe",
                    "v6.exe", "v7.exe", "v8.exe", "v9.exe", "v10.exe"]
    getSource(mr_list, ts, num_of_samples)

    for mu in range(1, len(mutants_list)):
        MG = getMG(mu, ts, num_of_samples, mr_list)
        pf = getpf(mu, ts, num_of_samples, mr_list)
        MG, pf = getIndex(MG, pf)

        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue
        row = getMetrics(row, ws, mu, MG, pf)

    wb.save(path)

