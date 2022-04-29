'''
RQ3
Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
import numpy as np
import random
import os
import json
random.seed(1)
from DM import *
from publicFun import *
from openpyxl import load_workbook

def riskIndex(argv, dynamic):
    MGS = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = DeterMinant().Determinant(testcase[i], n)  # oracle
        result_s_m = dynamic.Determinant(testcase[i], n)
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    # p or f统计完

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    return MGS, Result


if __name__ == '__main__':
    path = '/Applications/work/data/MT/FAILTIM/Result/RQ3_3.xlsx'
    wb = load_workbook(path)
    string = 'determinant'
    del wb[string]
    ws = wb.create_sheet(string)
    row = 1
    n = 3
    MGSet = []
    pfset = []
    for mu in range(1, 8):
        # dynamic = DeterFactory("Mutant" + str(mu)).getDeter()
        # 读测试用例
        # with open('Result/Determinant/mutant' + str(mu) + '_rq1.json', 'r') as load_f:
        #     data = json.load(load_f)
        # data = json.loads(data)
        # source_case_set = data['source_case_set']
        #
        # pf, MG, Index, Risk = getRisk(source_case_set)
        # data = {
        #         'pf': pf, 'MG':MG, 'Risk':Risk, 'Index':Index
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('Result/Determinant/mutant' + str(mu) + '_rq3.json', 'w') as f:
        #     json.dump(json_str, f)
        # 读数据
        with open('/Applications/work/data/MT/FAILTIM/Result/' + string + '/mutant' + str(mu) + '_rq1_new.json', 'r') as load_f:
            data = json.load(load_f)
        data = json.loads(data)
        pf = data['pf']
        MG = data['MG']
        pfset.append(pf)
        MGSet.append(MG)
        if len(MG) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics_rq3_3(row, ws, mu, MG, pf)

    wb.save(path)
