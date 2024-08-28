from TSQ import *
import math
import numpy as np
import random
import json
from PublicFun import *
from openpyxl import load_workbook
random.seed(1)


def getSource(dynamic):
    Follow = []
    source_case_set = []
    while 1:
        a = random.randint(1, 10)
        b = random.randint(1, 10)
        c = random.randint(1, 10)
        if a >= b + c or b >= a + c or c >= a + b:
            continue
        else:
            triangle = [a, b, c]
            source_case_set.append(triangle)
            if len(source_case_set) >= 100:
                break

    for i in range(len(source_case_set)):
        MGS, follow_case = MTG(source_case_set[i], dynamic)
        Follow.append(follow_case)

    for i in range(len(Follow)):
        for j in range(len(Follow[i])):
            source_case_set.append(Follow[i][j])
    MG = []
    for i in range(len(source_case_set)):
        MGS, Result = riskIndex(source_case_set[i], dynamic)
        MG.append(MGS)
    source_case_set = modifycase_fairness(source_case_set, MG)

    if len(source_case_set) == 0:
        print('测试用例集不满足要求')

    return source_case_set


def riskIndex(argv, dynamic):
    MGS = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = Trisquare().trisquare(testcase[i])  # oracle
        result_s_m = dynamic.trisquare(testcase[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)

    return MGS, Result


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/result1.xlsx'
    wb = load_workbook(path)
    string = 'TSQ'
    del wb[string]
    ws = wb.create_sheet(string)
    source = []
    for mu in range(1, 7):
        dynamic = TriFactory("Mutant" + str(mu)).getTri()
        source_case_set = getSource(dynamic)
        MG_set = []
        pf_set = []
        for i in range(len(source_case_set)):
            MG, pf = riskIndex(source_case_set[i], dynamic)
            MG_set.append(MG)
            pf_set.append(pf)

        if len(MG_set) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics(row, ws, mu, MG_set, pf_set)

    wb.save(path)
