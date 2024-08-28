import random
from DM import *
from PublicFun import *
from openpyxl import load_workbook
random.seed(1)


def riskIndex(argv, dynamic):
    MGS = []
    # index = []
    Result = []
    testcase = []
    source_case = argv.copy()
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = DeterMinant().Determinant(testcase[i], n)  # oracle
        result_s_m = dynamic.Determinant(testcase[i], n)
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)

    return MGS, Result


def getSource(dynamic):
    Follow = []
    source_case_set = [[2, 7, 8, 1, 2, 7, 8, 8, 1], [7, 4, 6, 8, 4, 0, 5, 0, 0], [0, 4, 6, 8, 0, 0, 5, 0, 0],
                       [2, 0, 1, 3, 8, 2, 6, 4, 9], [2, 0, 0, 1, 5, 0, 2, 4, 6], [0, 0, 0, 1, 0, 0, 2, 4, 0]]

    for i in range(len(source_case_set)):
        MGS, follow_case = MTG(source_case_set[i], dynamic)
        Follow.append(follow_case)

    for i in range(len(Follow)):
        for j in range(len(Follow[i])):
            source_case_set.append(Follow[i][j])
    MG = []
    for i in range(len(source_case_set)):
        MGS, Result = riskIndex(source_case_set[i], dynamic)
        MG.append(MGS)
    source_case_set = modifycase_fairness(source_case_set, MG)

    if len(source_case_set) == 0:
        print('测试用例集不满足要求')

    return source_case_set


if __name__ == '__main__':
    row = 1
    path = '/Applications/work/data/MT/FAILTIM/Result/result1.xlsx'
    wb = load_workbook(path)
    string = 'DM'
    del wb[string]
    ws = wb.create_sheet(string)
    n = 3
    for mu in range(1, 8):
        dynamic = DeterFactory("Mutant" + str(mu)).getDeter()
        source_case_set = getSource(dynamic)
        MG_set = []
        pf_set = []
        for i in range(len(source_case_set)):
            MG, pf = riskIndex(source_case_set[i], dynamic)
            MG_set.append(MG)
            pf_set.append(pf)

        if len(MG_set) == 0:
            print("Mutant" + str(mu) + '不符合要求')
            continue

        row = getMetrics(row, ws, mu, MG_set, pf_set)

    wb.save(path)
