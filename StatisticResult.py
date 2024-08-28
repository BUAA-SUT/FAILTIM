from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import ticker
import scipy.stats as stats
import json

Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'dice',
           'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
           'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3',
           'Arithmetic Mean', 'Cohen', 'Fleiss']

path = '/Applications/work/data/MT/FAILTIM/Result/correlation_with.xlsx'
wb = load_workbook(path)
path2 = '/Applications/work/data/MT/FAILTIM/Result/correlation_without.xlsx'
wb2 = load_workbook(path2)
program = ['TSQ', 'DM', 'SMM', 'Tcas', 'KNN', 'MATH', 'PT',  'Num', 'DNA']  #
Data = {}
Datalist = []
Datafs = {}
Datalistfs = []
Datano = {}
Datalistno = []
for k in program:
    ws = wb[k]
    ws2 = wb2[k]
    max_row = ws.max_row
    max_col = ws.max_column
    datadist = {}
    Value = []
    datadistfs = {}
    Valuefs = []
    datadistno = {}
    Valueno = []
    for i in range(len(Formula)):
        Value.append([])
        Valuefs.append([])
        Valueno.append([])
    t = 0
    n = 2
    while True:
        value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
        Value[t].append(value)
        valuefs = ws.cell(n, max_col).value
        Valuefs[t].append(valuefs)
        valueno= ws2.cell(n, 2).value + ws2.cell(n, 4).value + ws2.cell(n, 6).value + 0.5 * ws2.cell(n, 3).value
        Valueno[t].append(valueno)
        t = t + 1
        n = n + 1
        if t == 30:
            t = 0
            n = n + 3
        if n > max_row:
            break
    for i in range(len(Formula)):
        # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
        #     continue
        data = {
            Formula[i]: Value[i]
        }
        datadist.update(data)
        datafs = {
            Formula[i]: Valuefs[i]
        }
        datadistfs.update(datafs)
        datano = {
            Formula[i]: Valueno[i]
        }
        datadistno.update(datano)
    Datalist.append(datadist)
    Datalistfs.append(datadistfs)
    Datalistno.append(datadistno)

for i in range(len(Formula)):
    # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
    #     continue
    dd = []
    value = []
    ddfs = []
    valuefs = []
    ddno = []
    valueno = []
    for j in range(len(program)):
        dd += Datalist[j][Formula[i]]
        ddfs += Datalistfs[j][Formula[i]]
        ddno += Datalistno[j][Formula[i]]
    data = {
        Formula[i]: dd
    }
    datafs = {
        Formula[i]: ddfs
    }
    datano = {
        Formula[i]: ddno
    }
    Data.update(data)
    Datafs.update(datafs)
    Datano.update(datano)

withfs = zip(Data.values(), Data.keys())
withfs = list(withfs)
withoutfs = zip(Datano.values(), Datano.keys())
withoutfs = list(withoutfs)
fs = zip(Datafs.values(), Datafs.keys())
fs = list(fs)

acc_with = []
acc_without = []
value_fs = []
for i in range(len(withfs[0][0])):
    vv = []
    vs = []
    vw = []
    for j in range(len(withfs)):
        vv.append(withfs[j][0][i])
        vw.append(withoutfs[j][0][i])
        vs.append(fs[j][0][i])
    acc_with.append(np.mean(vv))
    acc_without.append(np.mean(vw))
    value_fs.append(np.mean(vs) / 100)
improve = [(acc_without[i] - acc_with[i]) / acc_with[i] for i in range(0, len(acc_with))]

# 拟合曲线

points_tulpe = zip(value_fs, improve)
z_points_tulpe = list(zip(value_fs, improve))
z = sorted(points_tulpe)
z = list(z)
x = []
y = []
for i in range(len(z)):
    x.append(z[i][0])
    y.append(z[i][1])
# plt.figure()
parameter = np.polyfit(x, y, 3)
y2 = []
for i in x:
    y2.append(parameter[0] * i ** 3 + parameter[1] * i ** 2 + parameter[2] * i + parameter[3])
correlation = np.corrcoef(x, y)
p = np.poly1d(parameter)
# plt.plot(x, y2, color='g')
fig, ax = plt.subplots()
plt.scatter(x, y)
ax.plot(x, y2, color='g')
ax.xaxis.set_major_formatter(ticker.PercentFormatter(xmax=1))
plt.show()
plt.xlabel('Percentage of false satisfying MGs')
plt.ylabel('Improvement ratio of accuracy rate')
plt.show()


Data = {}
Datalist = []
for k in program:
    ws = wb[k]
    max_row = ws.max_row
    datadist = {}
    Value = []
    for i in range(len(Formula)):
        Value.append([])
    t = 0
    n = 2
    while True:
        value = ws.cell(n, 2).value + ws.cell(n, 4).value + ws.cell(n, 6).value + 0.5 * ws.cell(n, 3).value
        Value[t].append(value)
        t = t + 1
        n = n + 1
        if t == len(Formula):
            t = 0
            n = n + 3
        if n > max_row:
            break
    for i in range(len(Formula)):
        # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
        #     continue
        data = {
            Formula[i]: Value[i]
        }
        datadist.update(data)
    Datalist.append(datadist)


for i in range(len(Formula)):
    # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
    #     continue
    dd = []
    value = []
    for j in range(len(program)):
        dd += Datalist[j][Formula[i]]
    data = {
        Formula[i]: round(c, 2)
    }
    Data.update(data)

z = zip(Data.values(), Data.keys())
z = sorted(z, reverse=True)
values = []
for i in range(len(z)):
    values.append(z[i][0])
ave = round(sum(values)/len(values), 2)


path = '/Applications/work/data/MT/FAILTIM/Result/Bonferroni.xlsx'
wb = load_workbook(path)
string = 'Sheet1'
del wb[string]
ws = wb.create_sheet(string)
row = 1
datadist = {}
tablelist = {'Formulas': ['Mean', 'St.Dev.']}
datadist.update(tablelist)
mean = []
std = []
for i in range(len(Formula)):
    # if Formula[i] == 'Binary' or Formula[i] == 'Naish1':
    #     continue
    dd = []
    value = []
    for j in range(len(program)):
        dd += Datalist[j][Formula[i]]
    data = {
        Formula[i]: dd
    }
    Data.update(data)
    mean.append(round(sum(dd)/len(dd), 2))
    std.append(round(sum(dd)/len(dd), 2))

    value = [round(sum(dd)/len(dd), 2), round(sum(dd)/len(dd), 2)]
    data = {
        Formula[i]: value
    }
    datadist.update(data)

for i, j in datadist.items():  # i--公式名称, j--指标值
    ws.cell(row, 1).value = i  # 添加第 1 列的数据
    for col in range(2, len(j) + 2):  # values列表中索引
        ws.cell(row, col).value = j[col - 2]
    row += 1  # 行数

wb.save(path)
